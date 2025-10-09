from django.views.generic import CreateView, UpdateView, DeleteView, ListView
from .models import Participante,RegistroCorreo
import pandas as pd
import openpyxl
import qrcode
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.urls import reverse
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from PIL import Image, ImageDraw, ImageFont
from django.core.mail import EmailMessage
from django.conf import settings
from .utils import enviar_correo_participante
from django.db.models import Q
from django.utils import timezone
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import io
import os
from django.contrib.staticfiles import finders
import socket
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

def index(request):
    return render(request, "cliente/index.html")

@login_required
def formulario_clientes(request):
    # Tu lógica o vista de clientes
    return render(request, "cliente/lista.html")


 

def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # Se conecta a un servidor público solo para obtener tu IP local
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    finally:
        s.close()
    return ip

import pywhatkit
from io import BytesIO
from django.core.mail import EmailMultiAlternatives
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse
from email.mime.image import MIMEImage
import qrcode
import tempfile
import datetime
from twilio.rest import Client

def confirmar_pago(request, pk):
    participante = get_object_or_404(Participante, pk=pk)

    # Confirmar pago
    participante.pago_confirmado = True
    participante.save()

    # Generar el QR
    ip_local = get_local_ip()
    url = f"http://{ip_local}:8000{reverse('validar_entrada', args=[participante.token])}"
    qr_img = qrcode.make(url).convert("RGB")

    # Generar imagen final personalizada
    imagen_final = generar_imagen_personalizada(
        nombre_cliente=participante.nombres,
        paquete=participante.tipo_entrada,
        qr_img=qr_img
    )

    if imagen_final is None:
        messages.error(request, "❌ No se pudo generar la imagen de la entrada.")
        return redirect("participante_lista")

    # Guardar imagen en buffer
    buffer = BytesIO()
    imagen_final.save(buffer, format='PNG')
    buffer.seek(0)

    # Crear correo HTML
    asunto = "🎟️ Confirmación de tu entrada - El Despertar del Emprendedor"
    html_mensaje = f"""
    <html>
    <body>
        <p>Hola {participante.nombres},</p>
        <p>Gracias por tu compra. Adjunto encontrarás tu entrada personalizada
        para el evento <strong>El Despertar del Emprendedor</strong>.</p>
        <p>No olvides guardarla y mostrarla el día del evento.</p>
        <p>¡Nos vemos pronto!</p>
        <br>
        <img src="cid:entrada" alt="Entrada personalizada" style="max-width:100%; height:auto; display:block;">
    </body>
    </html>
    """

    email = EmailMultiAlternatives(
        subject=asunto,
        body="Tu correo no soporta HTML",  # texto plano
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[participante.correo],
    )
    email.attach_alternative(html_mensaje, "text/html")

    # Adjuntar imagen como inline
    img = MIMEImage(buffer.getvalue())
    img.add_header('Content-ID', '<entrada>')
    img.add_header('Content-Disposition', 'inline', filename='entrada.png')
    email.attach(img)

    # Enviar correo
    email.send()


    # --- WHATSAPP LOCAL ---
    # Guardar imagen en archivo temporal
    # Guardar imagen en archivo temporal para WhatsApp con tamaño optimizado
    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp_file:
        # Crear copia para WhatsApp y redimensionar para evitar pixelado
        imagen_para_whatsapp = imagen_final.copy()
        imagen_para_whatsapp.thumbnail((1080, 1440))  # Ajusta si quieres otra resolución
        imagen_para_whatsapp.save(tmp_file, format="JPEG", quality=95)
        tmp_path = tmp_file.name


    # Preparar número (solo dígitos)
    numero = f"+51{''.join(filter(str.isdigit, participante.celular))}"

    # Mensaj
    mensaje_whatsapp = f"Hola {participante.nombres}, esta es tu entrada para 'El Despertar del Emprendedor'."

    # Hora de envío (mínimo 1 minuto adelante)
    ahora = datetime.datetime.now()
    hora_envio = ahora.hour
    minuto_envio = ahora.minute + 1

    # Enviar imagen por WhatsApp Web
    pywhatkit.sendwhats_image(
        receiver=numero,
        img_path=tmp_path,
        caption=mensaje_whatsapp,
        wait_time=7,
        tab_close=True
    )
    # Registrar envío
    registro, created = RegistroCorreo.objects.get_or_create(
        participante=participante,
        defaults={"enviado": True, "fecha_envio": timezone.now()}
    )
    if not created:
        registro.enviado = True
        registro.fecha_envio = timezone.now()
        registro.save()

    messages.success(request, f"✅ Entrada enviada correctamente a {participante.correo}")
    return redirect("participante_lista")


def escalar_a_a4(imagen):
    # Tamaño A4 en píxeles a 300dpi (alta resolución)
    ancho_a4 = 2480  # 8.27 pulgadas * 300dpi
    alto_a4 = 3508   # 11.69 pulgadas * 300dpi

    ancho_img, alto_img = imagen.size

    # Escalar proporcionalmente
    factor = min(ancho_a4 / ancho_img, alto_a4 / alto_img)
    nuevo_ancho = int(ancho_img * factor)
    nuevo_alto = int(alto_img * factor)

    imagen_redimensionada = imagen.resize((nuevo_ancho, nuevo_alto), Image.ANTIALIAS)
    return imagen_redimensionada




import qrcode
from PIL import Image, ImageDraw, ImageFont
from django.contrib.staticfiles import finders

def generar_imagen_personalizada(nombre_cliente, qr_img=None, paquete=None):
    """
    Genera la imagen final compuesta con textos y QR.
    Parámetros:
        nombre_cliente (str)
        paquete (str|None) : texto del paquete (ej. "FULL ACCESS"). Si es None no se dibuja.
    Retorna:
        PIL.Image
    """

    # --- Buscar las imágenes base ---
    partes = []
    for i in range(1, 8):
        ruta = finders.find(f'img/parte0{i}.jpg')
        if ruta:
            partes.append(ruta)

    if not partes:
        raise ValueError("No se encontraron imágenes para generar la entrada.")

    imagenes = []

    for p in partes:
        img = Image.open(p).convert("RGB")  # asegurar modo RGB
        filename = p.replace("\\", "/").lower()  # para comparar nombres (portable)

        # --- parte02: texto principal ---
        if 'parte02.jpg' in filename:
            draw = ImageDraw.Draw(img)
            try:
                font_black_path = "C:/Windows/Fonts/ariblk.ttf"   # Arial Black (Windows)
                font_regular_path = "C:/Windows/Fonts/arial.ttf"  # Arial regular
                # tamaños basados en la altura para que queden proporcionados
                font_title = ImageFont.truetype(font_black_path, size=int(img.height * 0.05))
                font_name  = ImageFont.truetype(font_black_path, size=int(img.height * 0.06))
                font_body  = ImageFont.truetype(font_regular_path, size=int(img.height * 0.04))
            except Exception as e:
                print("⚠️ Error cargando fuentes:", e)
                font_title = font_name = font_body = ImageFont.load_default()

            # helper: texto con borde
            def draw_text_outline(draw_obj, pos, text, font, fill=(255,255,255), outline_fill=(0,0,0), outline_w=2):
                x, y = pos
                for dx in range(-outline_w, outline_w+1):
                    for dy in range(-outline_w, outline_w+1):
                        if dx != 0 or dy != 0:
                            draw_obj.text((x+dx, y+dy), text, font=font, fill=outline_fill)
                draw_obj.text((x, y), text, font=font, fill=fill)

            # helper: centrar texto horizontalmente
            def draw_centered(draw_obj, y, text, font, fill=(255,255,255)):
                ancho_texto = draw_obj.textlength(text, font=font)
                x = (img.width - ancho_texto) / 2
                draw_text_outline(draw_obj, (x, y), text, font, fill)

            # Título y bloque de texto
            draw_centered(draw, int(img.height * 0.085), "TU ENTRADA A EL DESPERTAR DEL EMPRENDEDOR", font_title)

            y_text = int(img.height * 0.25)
            lineas = [
                f"Hola {nombre_cliente},",
                "",
                "Gracias por unirte a EL DESPERTAR DEL EMPRENDEDOR",
                "",
                "Adjunto tu entrada personalizada:",
                "",
                "No olvides guardarla y mostrarla el",
                "     día del evento"
            ]
            for i_linea, linea in enumerate(lineas):
                if i_linea == 0:
                    draw_centered(draw, y_text, linea, font_name)
                    y_text += int(img.height * 0.09)
                else:
                    draw_centered(draw, y_text, linea, font_body)
                    y_text += int(img.height * 0.06)
    
        # --- parte03: pegar QR y texto del paquete ---
        if 'parte03.jpg' in filename:
            img = img.convert("RGB") 
            draw = ImageDraw.Draw(img)

            # Si no se pasa qr_img, generar uno por defecto
            if qr_img is None:
                url = f"https://miapp.com/validar/{nombre_cliente}"
                qr_img = qrcode.make(url).convert("RGB")


            # Coordenadas del cuadro blanco
            x1, y1, x2, y2 = 645, 218, 943, 509
            rect_w = x2 - x1
            rect_h = y2 - y1

            # margen interior
            margin = 10
            qr_size = min(rect_w, rect_h) - 2 * margin
            if qr_size <= 0:
                qr_size = max(50, int(img.width / 3))

            qr_img = qr_img.resize((qr_size, qr_size))

            # centrar QR
            qr_x = x1 + (rect_w - qr_size) // 2
            qr_y = y1 + (rect_h - qr_size) // 2
            img.paste(qr_img, (qr_x, qr_y))

            # -------------------------
            # --- Texto en la parte superior ---
            # -------------------------

            try:
                # Intentamos cargar una fuente personalizada (Arial Black) desde el sistema
                font_pkg_path = "C:/Windows/Fonts/ariblk.ttf"
                font_pkg = ImageFont.truetype(font_pkg_path, size=max(16, int(qr_size * 0.14)))
            except:
                # Si no se puede cargar la fuente, usamos la fuente predeterminada de PIL
                font_pkg = ImageFont.load_default()


            # Función para dibujar texto con contorno
            def draw_text_outline(draw_obj, pos, text, font, fill=(255, 255, 255),
                                outline_fill=(0, 0, 0), outline_w=3):
                """
                Dibuja un texto con contorno para mejorar visibilidad sobre cualquier fondo.
                """
                x, y = pos
                for dx in range(-outline_w, outline_w + 1):
                    for dy in range(-outline_w, outline_w + 1):
                        if dx != 0 or dy != 0:
                            draw_obj.text((x + dx, y + dy), text, font=font, fill=outline_fill)
                draw_obj.text(pos, text, font=font, fill=fill)


            # -------------------------
            # Añadir texto en la parte superior de la imagen
            # -------------------------
            print("! Paquete recibido:", paquete) #DEBUG

            if paquete:
                # Texto 1
                texto_arriba = f"Según tu paquete {paquete.upper()},"
                text_w = draw.textlength(texto_arriba, font=font_pkg)
                text_x = (img.width - text_w) / 2
                text_y = 10  # margen superior fijo (10 píxeles desde el borde superior)
                draw_text_outline(draw, (text_x, text_y), texto_arriba, font_pkg,
                                    fill=(255, 255, 255), outline_fill=(0, 0, 0), outline_w=2)

                # Texto 2, debajo del texto 1
                texto_arriba2 = "aquí tienes las indicaciones específicas"
                text_w2 = draw.textlength(texto_arriba2, font=font_pkg)
                text_x2 = (img.width - text_w2) / 2
                text_y2 = text_y + int(qr_size * 0.22)  # separación vertical del primer texto
                draw_text_outline(draw, (text_x2, text_y2), texto_arriba2, font_pkg,
                                        fill=(255, 255, 255), outline_fill=(0, 0, 0), outline_w=2)

                # Texto debajo del QR (sin cambios)
                texto_abajo = f"ENTRADA {paquete.upper()}"
                text_w3 = draw.textlength(texto_abajo, font=font_pkg)
                text_x3 = (img.width - text_w3) / 2
                text_y3 = qr_y + qr_size + 80  # posición debajo del QR
                draw_text_outline(draw, (text_x3, text_y3), texto_abajo, font_pkg,
                                fill=(0, 0, 0), outline_fill=(255, 255, 255), outline_w=2)


                # Añadir la imagen procesada a la lista
        imagenes.append(img)



    # --- unir todas las partes ---
    if not imagenes:
        raise ValueError("No se generaron imágenes (lista vacía).")
    # === UNIR TODAS LAS PARTES EN UNA SOLA IMAGEN ===
    ancho = max(img.width for img in imagenes)
    alto_total = sum(img.height for img in imagenes)
    imagen_final = Image.new('RGB', (ancho, alto_total), (255, 255, 255))

    y_offset = 0
    for img in imagenes:
        imagen_final.paste(img, (0, y_offset))
        y_offset += img.height

    return imagen_final


def limpiar_tipo_entrada(valor):
    if not isinstance(valor, str):
        return "EMPRENDEDOR"  # valor por defecto si es vacío
    # Extrae la parte después del guion
    tipo = valor.split('-')[-1].strip().upper()
    # Validar que sea uno de los permitidos
    if tipo not in ["FULL ACCES", "EMPRESARIAL", "EMPRENDEDOR"]:
        tipo = "EMPRENDEDOR"  # default
    return tipo






class ParticipanteCreateView(CreateView):
    model = Participante
    fields = ['nombres','apellidos','dni','celular','correo','tipo_entrada','cantidad']
    template_name = 'cliente/participante_form.html'
    success_url = reverse_lazy('participante_lista')


class ParticipanteUpdateView(UpdateView):
    model = Participante
    fields = ['nombres','apellidos','dni','celular','correo','tipo_entrada','cantidad','precio']
    template_name = 'cliente/participante_form.html'
    success_url = reverse_lazy('participante_lista')

class ParticipanteDeleteView(DeleteView):
    model = Participante
    template_name = 'cliente/participante_confirm_delete.html'
    success_url = reverse_lazy('participante_lista')
 
class ParticipanteListView(ListView):
    model = Participante
    template_name = 'cliente/lista.html'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get("q")  # Obtiene el valor del input de búsqueda
        if q:
            queryset = queryset.filter(
                Q(nombres__icontains=q) | Q(dni__icontains=q)
            )
        return queryset



from django.shortcuts import redirect
from django.contrib import messages
import pandas as pd
from .models import Participante

def limpiar_tipo_entrada(valor):
    if pd.isna(valor):
        return "EMPRENDEDOR"
    # Tomar solo la parte después del guion y convertir a mayúscula
    return valor.split('-')[-1].strip().upper()

def valor_seguro(x):
    if pd.isna(x) or x == 0:
        return ""
    return str(x).strip()


def importar_excel(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        archivo = request.FILES['excel_file']
        try:
            # Leer Excel
            df = pd.read_excel(archivo)
            print(df.columns)  # Verifica los nombres de columnas

            # Asegúrate de que los nombres coincidan
            df = df.rename(columns={
                'Nombre': 'Nombre',
                'DNI': 'DNI',
                'TELEFONO': 'TELEFONO',
                'Correo electrónico': 'Correo',
                'Tipo de entrada': 'Tipo_Entrada'
            })

            # Extraer solo la parte después del guion
            df['Tipo_Entrada'] = df['Tipo_Entrada'].astype(str).apply(lambda x: x.split('-')[-1].strip())

            # Iterar y crear participantes
            for _, row in df.iterrows():
                if pd.isna(row['DNI']) or pd.isna(row['Nombre']):
                    continue  # Ignora filas vacías críticas

                Participante.objects.create(
                    nombres=row['Nombre'],
                    apellidos="",  # Puedes separar apellido si quieres
                    dni=str(row['DNI']),
                    celular=str(row['TELEFONO']) if not pd.isna(row['TELEFONO']) else '',
                    correo=row['Correo'] if not pd.isna(row['Correo']) else '',
                    tipo_entrada=row['Tipo_Entrada'],
                    cantidad=1
                )

            messages.success(request, "✅ Participantes importados correctamente.")
        except Exception as e:
            messages.error(request, f"❌ Error al importar Excel: {e}")

    return redirect('participante_lista')

def generar_qr(request, token):
    """
    Genera un PNG con un QR que apunta a la vista 'validar_entrada' usando el token del participante.
    """

    # 1) Recuperar el participante o devolver 404 si no existe
    participante = get_object_or_404(Participante, token=token)

    # ---------- Construcción de la URL que queremos codificar ----------
    # Opción A: Forzar host con tu IP local (útil para pruebas cuando tu servidor corre en otra máquina)
    # 🔹 Construir la URL automáticamente (sin IP fija)
    ip_local = get_local_ip()
    url = f"http://{ip_local}:8000{reverse('validar_entrada', args=[participante.token])}"

    # Opción B (recomendada si el host actual es el correcto):
    # url = request.build_absolute_uri(reverse('validar_entrada', args=[participante.token]))
    # esto genera automáticamente "http(s)://<host>/ruta" usando request.scheme y request.get_host()

    print("👉 URL del QR generado:", url)  # solo para desarrollo; en producción usa logging

    # ---------- Crear el QR con qrcode.QRCode (más control que qrcode.make) ----------
    qr_obj = qrcode.QRCode(
        version=None,  # None -> la librería calcula el tamaño necesario automáticamente
        error_correction=qrcode.constants.ERROR_CORRECT_M,  # tolerancia a errores (M es una buena elección)
        box_size=10,   # tamaño de cada "cuadro" del QR en pixeles
        border=4       # borde blanco (mínimo 4 recomendado por especificación)
    )
    qr_obj.add_data(url)
    qr_obj.make(fit=True)

    img = qr_obj.make_image(fill_color="black", back_color="white").convert("RGB")

    # ---------- Devolver la imagen como respuesta PNG ----------
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="image/png")
    # Si quieres que el navegador muestre el QR en línea:
    response['Content-Disposition'] = 'inline; filename="qr.png"'
    # Si deseas que lo descargue:
    # response['Content-Disposition'] = 'attachment; filename="qr.png"'

    return response



def mostrar_qr(request, pk):
    participante = get_object_or_404(Participante, pk=pk)

    # Ruta de la imagen de fondo
    fondo_path = os.path.join(settings.BASE_DIR, 'cliente', 'static', 'cliente', 'img', 'exponentes.png')
    img = Image.open(fondo_path).convert("RGB")
    draw = ImageDraw.Draw(img)

    # Escribir los datos encima
    font = ImageFont.truetype("arial.ttf", 24)
    draw.text((20, 50), f"Código: {participante.cod_cliente}", fill="black", font=font)
    draw.text((20, 100), f"DNI: {participante.dni}", fill="black", font=font)

    # Respuesta como imagen
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response

def validar_entrada(request, token):
    participante = get_object_or_404(Participante, token=token)

    if not participante.usado:
        participante.usado = True
        participante.save()
        return render(request, "cliente/entrada_valida.html", {"participante": participante})
    else:
        return render(request, "cliente/entrada_usada.html", {"participante": participante})


def marcar_ingreso(request, pk):
    participante = get_object_or_404(Participante, pk=pk)
    if not participante.entrada_usada:  # solo marcar si aún no entró
        participante.entrada_usada = True
        participante.save()
    return redirect('lista')  # Ajusta al nombre de tu lista

def exportar_excel(request):
    # Obtener datos de los participantes
    participantes = Participante.objects.all().values()
    if not participantes:
        return HttpResponse("No hay participantes para exportar.", content_type="text/plain")

    df = pd.DataFrame(participantes)

    # Crear buffer en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Participantes')

        workbook = writer.book
        worksheet = writer.sheets['Participantes']

        # 🎨 Estilos personalizados (morado y negro)
        header_font = Font(bold=True, color="FFFFFF")  # blanco
        header_fill = PatternFill(start_color="6A0DAD", end_color="6A0DAD", fill_type="solid")  # morado oscuro
        cell_font = Font(color="000000")  # texto negro
        center_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # ✅ Aplicar estilo a los encabezados
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border_style

        # ✅ Aplicar estilo a las filas de datos
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.font = cell_font
                cell.border = border_style
                cell.alignment = Alignment(vertical="center")

        # ✅ Ajustar ancho automático de columnas
        for column_cells in worksheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            worksheet.column_dimensions[column].width = max_length + 2

        # ✅ Agregar una franja decorativa con el título
        worksheet.insert_rows(1)
        worksheet.merge_cells('A1:{}1'.format(get_column_letter(worksheet.max_column)))
        titulo = worksheet.cell(row=1, column=1)
        titulo.value = "🎟️ Lista de Participantes - Exportación"
        titulo.font = Font(bold=True, size=14, color="FFFFFF")
        titulo.fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")  # morado intenso
        titulo.alignment = Alignment(horizontal="center", vertical="center")

        # ✅ Agregar fecha de exportación
        fila_fecha = worksheet.max_row + 2
        worksheet.merge_cells(f"A{fila_fecha}:C{fila_fecha}")
        fecha_cell = worksheet.cell(row=fila_fecha, column=1)
        fecha_cell.value = f"📅 Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        fecha_cell.font = Font(italic=True, color="555555")
        fecha_cell.alignment = Alignment(horizontal="left")

    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Participantes_MoradoNegro.xlsx'
    response.write(output.getvalue())
    return response

def panel_control(request):
    registros = RegistroCorreo.objects.all().order_by('-fecha_envio')
    return render(request, 'cliente/panel_control.html', {'registros': registros})

def reenviar_correo(request, pk):
    registro = get_object_or_404(RegistroCorreo, pk=pk)
    participante = registro.participante

    # Reenviar correo
    enviar_correo_participante(participante)

    # Actualizar registro
    registro.enviado = True
    registro.fecha_envio = timezone.now()
    registro.save()

    return redirect("panel_control")

def exportar_excel_control(request):
    registros = RegistroCorreo.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Encabezados
    columnas = ["Nombres", "Correo", "Tipo Entrada", "Fecha Envío", "Enviado"]
    ws.append(columnas)

    # Datos
    for r in registros:
        ws.append([
            f"{r.participante.nombres} {r.participante.apellidos}",
            r.participante.correo,
            r.participante.tipo_entrada,
            r.fecha_envio.strftime("%d/%m/%Y %H:%M") if r.fecha_envio else "",
            "Sí" if r.enviado else "No"
        ])

    # Respuesta HTTP con archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="control.xlsx"'
    wb.save(response)
    return response

def registros_json(request):
    registros = RegistroCorreo.objects.select_related("participante").order_by("-fecha_envio")
    data = {
        "registros": [
            {
                "id": r.id,
                "nombre": f"{r.participante.nombres} {r.participante.apellidos}",
                "correo": r.participante.correo,
                "entrada": r.participante.tipo_entrada,
                "fecha": r.fecha_envio.strftime("%d/%m/%Y") if r.fecha_envio else "",
                "estado": r.enviado,
            }
            for r in registros
        ]
    }
    return JsonResponse(data)
 


def exportar_pdf_control(request):
    registros = RegistroCorreo.objects.all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Encabezado
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Reporte de Registros", styles["Heading1"]))

    # Tabla
    data = [["Nombres", "Correo", "Tipo Entrada", "Fecha Envío", "Enviado"]]
    for r in registros:
        data.append([
            f"{r.participante.nombres} {r.participante.apellidos}",
            r.participante.correo,
            r.participante.tipo_entrada,
            r.fecha_envio.strftime("%d/%m/%Y") if r.fecha_envio else "",
            "Sí" if r.enviado else "No"
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.gray),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)

    # Generar PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="control.pdf"'
    response.write(pdf)
    return response


from django.shortcuts import render
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import base64
from django.contrib.staticfiles import finders

def preview_imagen_final(request):
    """
    Vista que genera una imagen final compuesta por varias imágenes de fondo,
    personalizando una de ellas (parte02.jpg) con el nombre del cliente y un texto
    adicional. El resultado se devuelve como imagen en base64 para previsualizarla
    en el navegador sin necesidad de guardarla físicamente en disco.
    """

    # 1. Obtener el nombre del cliente desde la URL (ej: ?nombre=Leonardo)
    # Si no se pasa en la URL, se usa "Nombre Cliente" como valor por defecto.
    nombre_cliente = request.GET.get('nombre', 'Nombre Cliente')

    # 2. Buscar las imágenes parte01.jpg, parte02.jpg, ..., parte07.jpg en static/img
    partes = []
    for i in range(1, 8):
        ruta = finders.find (f'img/parte0{i}.jpg')  # localiza archivo estático
        if ruta:
            partes.append(ruta)

    # 3. Validación: si no hay imágenes encontradas, retornar error 404
    if not partes:
        return HttpResponse("No se encontraron imágenes para preview.", status=404)

    # Lista donde se guardarán las imágenes abiertas
    imagenes = []

    # 4. Procesar cada imagen encontrada
    for p in partes:
        img = Image.open(p)  # abrir la imagen con Pillow (PIL)

        # Si es la parte02.jpg, se le dibuja el texto personalizado
        if 'parte02.jpg' in p:
            draw = ImageDraw.Draw(img)

            try:
                # Cargar fuentes (asegúrate de que existan en static/fonts)
                font_path_bold = finders.find('fonts/arialbd.ttf')
                font_path_regular = finders.find('fonts/arial.ttf')

                # Calcular tamaño de letra dinámico según ancho de la imagen
                ancho_img = img.width
                font_title = ImageFont.truetype(font_path_bold, size=int(ancho_img/15))
                font_body = ImageFont.truetype(font_path_regular, size=int(ancho_img/25))
            except:
                # Si no encuentra las fuentes, usar fuentes por defecto
                font_title = ImageFont.load_default()
                font_body = ImageFont.load_default()

            # Función auxiliar para texto con borde (para mejor visibilidad)
            def draw_text_outline(draw_obj, position, text, font, fill,
                                  outline_fill='black', outline_width=2):
                x, y = position
                for dx in range(-outline_width, outline_width+1):
                    for dy in range(-outline_width, outline_width+1):
                        if dx != 0 or dy != 0:
                            draw_obj.text((x+dx, y+dy), text, font=font, fill=outline_fill)
                draw_obj.text(position, text, font=font, fill=fill)

            # Función auxiliar para centrar texto horizontalmente
            def draw_centered(draw_obj, y, text, font, fill=(255, 255, 255)):
                ancho_texto = draw_obj.textlength(text, font=font)
                x = (img.width - ancho_texto) / 2
                draw_text_outline(draw_obj, (x, y), text, font, fill)

            # Escribir el título del evento en la parte superior
            draw_centered(draw, 70, "TU ENTRADA A EL DESPERTAR DEL EMPRENDEDOR", font_title)

            # Escribir cuerpo del mensaje (varias líneas)
            y_text = 100
            lineas = [
                f"Hola {nombre_cliente}",
                "Gracias por unirte a EL DESPERTAR DEL",
                "         EMPRENDEDOR",
                "",
                "Adjunto tu entrada personalizada:",
                "",
                "No olvides guardarla y mostrarla el",
                "     dia del evento"
            ]
            for linea in lineas:
                draw_centered(draw, y_text, linea, font_body)
                y_text += int(ancho_img / 25) + 10  # espacio entre líneas proporcional

        # Agregar imagen (modificada o no) a la lista final
        imagenes.append(img)

    # 5. Crear una imagen nueva que una todas las partes en vertical
    ancho = max(img.width for img in imagenes)              # ancho máximo
    alto_total = sum(img.height for img in imagenes)        # altura sumada
    imagen_final = Image.new('RGB', (ancho, alto_total), (255, 255, 255))

    # Pegar cada imagen en su posición dentro del lienzo final
    y_offset = 0
    for img in imagenes:
        imagen_final.paste(img, (0, y_offset))
        y_offset += img.height

    # 6. Convertir la imagen resultante a base64 para incrustarla en HTML
    buffer = BytesIO()
    imagen_final.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.read()).decode('utf-8')

    # 7. Retornar la imagen como <img src="data:..."> en la plantilla
    return render(request, 'cliente/preview_imagen.html', {'img_base64': img_base64})














