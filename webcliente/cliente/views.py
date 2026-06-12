import io
import os
import csv
import time
import logging
import base64
import secrets
import urllib.parse
import requests
import qrcode
from decimal import Decimal
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO

from django.conf import settings
from django.db.models import Q, Max, IntegerField, Sum
from django.db.models.functions import Cast, Substr
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import views as auth_views
from django.core.cache import cache
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import CreateView, UpdateView, DeleteView, ListView
from django.contrib.staticfiles import finders
from twilio.rest import Client
from email.mime.image import MIMEImage
from django.core.mail import EmailMultiAlternatives, EmailMessage
from django.core.mail.backends.smtp import EmailBackend
from django.core.files import File

from django.core.paginator import Paginator
from openpyxl.styles import Font, PatternFill

# Import models & forms
from .models import Evento, Tarifa, PerfilUsuario, Participante, Voucher, RegistroCorreo, Previaparticipantes
from .forms import ParticipanteForm

logger = logging.getLogger(__name__)

# ==========================================
# 🛡️ DECORADORES Y ACCESOS
# ==========================================
def rol_requerido(roles_permitidos):
    def decorator(view_func):
        @login_required(login_url='/participantes/login/')
        def _wrapped_view(request, *args, **kwargs):
            perfil = get_object_or_404(PerfilUsuario, user=request.user)
            if perfil.rol in roles_permitidos:
                return view_func(request, *args, **kwargs)
            messages.error(request, "Acceso restringido para tu tipo de rol.")
            return redirect('dashboard_eventos')
        return _wrapped_view
    return decorator


def verificar_permiso_evento(user, evento):
    perfil, _ = PerfilUsuario.objects.get_or_create(user=user, defaults={'rol': 'REGISTRADOR'})
    if perfil.rol == 'SUPERADMIN':
        return True
    return evento in perfil.eventos.all()


# ==========================================
# 🏠 REDIRECCIONES DE INICIO
# ==========================================
@login_required(login_url='/participantes/login/')
def home_redirect(request):
    # Los asesores de lotes entran directo al mapa
    if request.user.groups.filter(name='Asesores').exists():
        return redirect('plano')
    return redirect('dashboard_eventos')


# ==========================================
# 🔒 LOGIN CON PROTECCIÓN ANTI-FUERZA BRUTA
# Máx. 3 intentos por IP → bloqueo escalonado (15s, 30s, 60s, 120s... tope 15 min)
# ==========================================
LOGIN_MAX_INTENTOS = 3
LOGIN_ESPERA_BASE  = 15      # segundos del primer bloqueo
LOGIN_ESPERA_TOPE  = 900     # tope de espera (15 minutos)


def _client_ip(request):
    """IP real del cliente (considera el proxy de Render vía X-Forwarded-For)."""
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', 'desconocida')


class LoginConThrottle(auth_views.LoginView):
    template_name = 'cliente/login.html'

    def _keys(self):
        ip = _client_ip(self.request)
        return (f"login_fails_{ip}", f"login_lock_{ip}", f"login_level_{ip}")

    def _espera_restante(self):
        _, k_lock, _ = self._keys()
        restante = int(cache.get(k_lock, 0) - time.time())
        return restante if restante > 0 else 0

    def post(self, request, *args, **kwargs):
        # Si está bloqueado, ni siquiera intentamos autenticar
        restante = self._espera_restante()
        if restante > 0:
            messages.error(request, f"🔒 Demasiados intentos. Espera {restante} segundos antes de volver a intentar.")
            return self.render_to_response(self.get_context_data(form=self.get_form()))
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        # Login correcto → limpiar todos los contadores de esa IP
        cache.delete_many(list(self._keys()))
        return super().form_valid(form)

    def form_invalid(self, form):
        # Credenciales incorrectas → contar el intento
        k_fails, k_lock, k_level = self._keys()
        fails = cache.get(k_fails, 0) + 1
        if fails >= LOGIN_MAX_INTENTOS:
            # Bloqueo escalonado: 15s, 30s, 60s, 120s... (cada bloqueo dura el doble)
            level  = cache.get(k_level, 0) + 1
            espera = min(LOGIN_ESPERA_BASE * (2 ** (level - 1)), LOGIN_ESPERA_TOPE)
            cache.set(k_lock,  time.time() + espera, timeout=espera + 120)
            cache.set(k_level, level, timeout=86400)   # la escalada se recuerda 24 h
            cache.delete(k_fails)
            messages.error(self.request, f"🔒 Demasiados intentos fallidos. Acceso bloqueado por {espera} segundos.")
        else:
            cache.set(k_fails, fails, timeout=3600)
            restantes = LOGIN_MAX_INTENTOS - fails
            messages.error(self.request, f"Usuario o contraseña incorrectos. Te queda(n) {restantes} intento(s).")
        return super().form_invalid(form)


# ==========================================
# 🏢 DASHBOARD DE EVENTOS (SaaS GENERAL)
# ==========================================
@login_required(login_url='/participantes/login/')
def dashboard_eventos(request):
    perfil, created = PerfilUsuario.objects.get_or_create(
        user=request.user, 
        defaults={'rol': 'SUPERADMIN' if request.user.is_superuser else 'REGISTRADOR'}
    )
    
    if perfil.rol == 'SUPERADMIN':
        eventos = Evento.objects.all()
    else:
        eventos = perfil.eventos.all()
    
    # Calcular estadísticas dinámicas para cada evento
    for ev in eventos:
        ev.total_participantes = ev.participantes.count()
        ev.confirmados = ev.participantes.filter(pago_confirmado=True).count()
        ev.ingresos = ev.participantes.filter(pago_confirmado=True).aggregate(total=Sum('total_pagar'))['total'] or Decimal('0.00')
    
    return render(request, 'cliente/dashboard_eventos.html', {
        'eventos': eventos,
        'perfil': perfil
    })


# ==========================================
# ⚙️ CREACIÓN / EDICIÓN DE EVENTOS
# ==========================================
@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def evento_crear_editar(request, pk=None):
    perfil = get_object_or_404(PerfilUsuario, user=request.user)
    evento = None
    if pk:
        evento = get_object_or_404(Evento, pk=pk)
        if perfil.rol != 'SUPERADMIN' and evento not in perfil.eventos.all():
            messages.error(request, "No tienes autorización para editar este evento.")
            return redirect('dashboard_eventos')

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        fecha = request.POST.get("fecha_evento")
        
        # Límites
        aforo_maximo = int(request.POST.get("aforo_maximo", 500) or 500)
        limite_entradas_persona = int(request.POST.get("limite_entradas_persona", 5) or 5)

        # SMTP (campos legacy — ya no se usan en el formulario, se mantienen por compatibilidad)
        smtp_host = request.POST.get("smtp_host") or "smtp.sendgrid.net"
        smtp_port = request.POST.get("smtp_port") or 587
        smtp_user = request.POST.get("smtp_user") or "apikey"
        smtp_password = request.POST.get("smtp_password") or None
        default_from_email = request.POST.get("default_from_email") or settings.DEFAULT_FROM_EMAIL
        
        # WhatsApp Provider configuration
        whatsapp_provider = request.POST.get("whatsapp_provider", "INACTIVE")
        whatsapp_api_url = request.POST.get("whatsapp_api_url", "")
        whatsapp_api_headers = request.POST.get("whatsapp_api_headers", "")
        whatsapp_api_payload = request.POST.get("whatsapp_api_payload", "")
        
        # Twilio & ImgBB
        twilio_account_sid = request.POST.get("twilio_account_sid", "")
        twilio_auth_token = request.POST.get("twilio_auth_token", "")
        twilio_phone_number = request.POST.get("twilio_phone_number", "")
        twilio_whatsapp_number = request.POST.get("twilio_whatsapp_number", "")
        imgbb_api_key = request.POST.get("imgbb_api_key", "")
        
        # Estética
        color_primario = request.POST.get("color_primario", "#7b1fa2")

        if not evento:
            evento = Evento.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                fecha_evento=fecha or None,
                aforo_maximo=aforo_maximo,
                limite_entradas_persona=limite_entradas_persona,
                smtp_host=smtp_host,
                smtp_port=smtp_port or 587,
                smtp_user=smtp_user,
                smtp_password=smtp_password,
                default_from_email=default_from_email,
                whatsapp_provider=whatsapp_provider,
                whatsapp_api_url=whatsapp_api_url,
                whatsapp_api_headers=whatsapp_api_headers,
                whatsapp_api_payload=whatsapp_api_payload,
                twilio_account_sid=twilio_account_sid,
                twilio_auth_token=twilio_auth_token,
                twilio_phone_number=twilio_phone_number,
                twilio_whatsapp_number=twilio_whatsapp_number,
                imgbb_api_key=imgbb_api_key,
                color_primario=color_primario
            )
            perfil.eventos.add(evento)
            messages.success(request, f"¡Evento '{evento.nombre}' creado exitosamente!")
        else:
            evento.nombre = nombre
            evento.descripcion = descripcion
            if fecha:
                evento.fecha_evento = fecha
            evento.aforo_maximo = aforo_maximo
            evento.limite_entradas_persona = limite_entradas_persona
            evento.smtp_host = smtp_host
            evento.smtp_port = smtp_port or 587
            evento.smtp_user = smtp_user
            if smtp_password:
                evento.smtp_password = smtp_password
            evento.default_from_email = default_from_email
            evento.whatsapp_provider = whatsapp_provider
            evento.whatsapp_api_url = whatsapp_api_url
            evento.whatsapp_api_headers = whatsapp_api_headers
            evento.whatsapp_api_payload = whatsapp_api_payload
            evento.twilio_account_sid = twilio_account_sid
            evento.twilio_auth_token = twilio_auth_token
            evento.twilio_phone_number = twilio_phone_number
            evento.twilio_whatsapp_number = twilio_whatsapp_number
            evento.imgbb_api_key = imgbb_api_key
            evento.color_primario = color_primario
            
            # Subir imágenes
            if request.FILES.get("imagen_fondo"):
                evento.imagen_fondo = request.FILES.get("imagen_fondo")
            if request.FILES.get("logo"):
                evento.logo = request.FILES.get("logo")
            if request.FILES.get("banner"):
                evento.banner = request.FILES.get("banner")
                
            evento.save()
            messages.success(request, f"Evento '{evento.nombre}' actualizado.")

        # Guardar / Actualizar Tarifas Dinámicas del Evento
        tariff_ids = request.POST.getlist("tariff_id")
        tariff_names = request.POST.getlist("tariff_name")
        tariff_p1s = request.POST.getlist("tariff_p1")
        tariff_p2s = request.POST.getlist("tariff_p2")
        tariff_p3s = request.POST.getlist("tariff_p3")
        tariff_puertas = request.POST.getlist("tariff_puerta")

        saved_ids = []

        for i in range(len(tariff_names)):
            t_name = tariff_names[i].strip()
            if not t_name:
                continue
                
            t_id = tariff_ids[i] if i < len(tariff_ids) and tariff_ids[i] else None
            t_p1 = Decimal(tariff_p1s[i] or 0) if i < len(tariff_p1s) else Decimal(0)
            t_p2 = Decimal(tariff_p2s[i] or 0) if i < len(tariff_p2s) else Decimal(0)
            t_p3 = Decimal(tariff_p3s[i] or 0) if i < len(tariff_p3s) else Decimal(0)
            t_puerta = Decimal(tariff_puertas[i] or 0) if i < len(tariff_puertas) else Decimal(0)

            if t_id:
                t_obj = Tarifa.objects.filter(pk=t_id, evento=evento).first()
                if t_obj:
                    t_obj.tipo_entrada = t_name
                    t_obj.preventa_1 = t_p1
                    t_obj.preventa_2 = t_p2
                    t_obj.preventa_3 = t_p3
                    t_obj.puerta = t_puerta
                    t_obj.save()
                    saved_ids.append(t_obj.id)
            else:
                t_obj = Tarifa.objects.create(
                    evento=evento,
                    tipo_entrada=t_name,
                    preventa_1=t_p1,
                    preventa_2=t_p2,
                    preventa_3=t_p3,
                    puerta=t_puerta
                )
                saved_ids.append(t_obj.id)

        # Eliminar las tarifas que ya no estén presentes en el formulario
        Tarifa.objects.filter(evento=evento).exclude(id__in=saved_ids).delete()

        return redirect('dashboard_eventos')

    tarifas = None
    if evento:
        tarifas = Tarifa.objects.filter(evento=evento)

    return render(request, 'cliente/evento_form.html', {
        'evento': evento,
        'tarifas': tarifas
    })


# ==========================================
# 🗑️ ELIMINAR EVENTO
# ==========================================
@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def evento_eliminar(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    perfil = get_object_or_404(PerfilUsuario, user=request.user)

    if perfil.rol != 'SUPERADMIN' and evento not in perfil.eventos.all():
        messages.error(request, "No tienes autorización para eliminar este evento.")
        return redirect('dashboard_eventos')

    # Solo permitir borrado via POST (protección CSRF + accidentalidad)
    if request.method != 'POST':
        messages.error(request, "Acción no permitida.")
        return redirect('dashboard_eventos')

    nombre = evento.nombre

    # Eliminar primero todos los participantes del evento (y sus tarifas/vouchers)
    # para evitar ProtectedError con cualquier relación que pueda quedar
    evento.participantes.all().delete()
    evento.previa_participantes.all().delete()

    # Borrar archivos de media asociados al evento
    for campo in [evento.imagen_fondo, evento.logo, evento.banner]:
        if campo:
            try:
                campo.delete(save=False)
            except Exception:
                pass

    evento.delete()
    messages.success(request, f"Evento '{nombre}' y todos sus datos fueron eliminados permanentemente.")
    return redirect('dashboard_eventos')


# ══════════════════════════════════════════════════════════════════════
# 📧 CAPA 1 — SMTP interno (helper privado)
# ══════════════════════════════════════════════════════════════════════
def _enviar_via_smtp(destinatarios, asunto, html_mensaje, cuerpo_texto, imagen_final_buffer,
                     host, port, user, password, from_email):
    """Envío SMTP puro. Usado como fallback cuando no hay OAuth2."""
    use_tls = getattr(settings, 'EMAIL_USE_TLS', True)
    use_ssl  = getattr(settings, 'EMAIL_USE_SSL', False)
    try:
        if getattr(settings, 'EMAIL_BACKEND', '') == 'django.core.mail.backends.locmem.EmailBackend':
            from django.core.mail import get_connection
            backend = get_connection()
        else:
            backend = EmailBackend(
                host=host, port=port,
                username=user, password=password,
                use_tls=use_tls, use_ssl=use_ssl,
                fail_silently=False
            )
        email_msg = EmailMultiAlternatives(
            subject=asunto,
            body=cuerpo_texto or "Tu cliente de correo no soporta mensajes en formato HTML.",
            from_email=from_email,
            to=destinatarios,
            connection=backend
        )
        email_msg.attach_alternative(html_mensaje, "text/html")
        if imagen_final_buffer:
            imagen_final_buffer.seek(0)
            img = MIMEImage(imagen_final_buffer.read())
            img.add_header('Content-ID', '<entrada>')
            img.add_header('Content-Disposition', 'inline', filename='entrada.png')
            email_msg.attach(img)
        email_msg.send()
        logger.info(f"📧 SMTP: correo enviado a {destinatarios} desde {from_email} via {host}:{port}")
        return True
    except Exception as e:
        logger.error(f"❌ SMTP ({host}:{port}, user={user}): {e}", exc_info=True)
        return False


# ══════════════════════════════════════════════════════════════════════
# 📧 CAPA 2 — Gmail API con OAuth2 (sin SMTP, sin contraseñas)
# ══════════════════════════════════════════════════════════════════════
def enviar_con_gmail_api(destinatarios, asunto, html_mensaje, imagen_final_buffer=None,
                         from_email=None, refresh_token=None):
    """Envía correo usando Gmail API + OAuth2. No requiere App Password."""
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage as StdMIMEImage

    if not isinstance(destinatarios, list):
        destinatarios = [destinatarios]

    try:
        # 1. Obtener access_token fresco con el refresh_token
        token_resp = requests.post(
            'https://oauth2.googleapis.com/token',
            data={
                'client_id':     settings.GOOGLE_CLIENT_ID,
                'client_secret': settings.GOOGLE_CLIENT_SECRET,
                'refresh_token': refresh_token,
                'grant_type':    'refresh_token',
            },
            timeout=15
        )
        token_data   = token_resp.json()
        access_token = token_data.get('access_token')

        if not access_token:
            logger.error(f"Gmail API: no se pudo renovar token: {token_data}")
            return False

        # 2. Construir mensaje MIME
        msg = MIMEMultipart('related')
        msg['to']      = ', '.join(destinatarios)
        msg['from']    = from_email
        msg['subject'] = asunto

        alt = MIMEMultipart('alternative')
        alt.attach(MIMEText("Tu cliente de correo no soporta HTML.", 'plain', 'utf-8'))
        alt.attach(MIMEText(html_mensaje, 'html', 'utf-8'))
        msg.attach(alt)

        if imagen_final_buffer:
            imagen_final_buffer.seek(0)
            img = StdMIMEImage(imagen_final_buffer.read())
            img.add_header('Content-ID', '<entrada>')
            img.add_header('Content-Disposition', 'inline', filename='entrada.png')
            msg.attach(img)

        # 3. Enviar via Gmail API REST
        raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode('utf-8')
        send_resp = requests.post(
            'https://gmail.googleapis.com/gmail/v1/users/me/messages/send',
            headers={
                'Authorization': f'Bearer {access_token}',
                'Content-Type':  'application/json',
            },
            json={'raw': raw_message},
            timeout=30
        )

        if send_resp.status_code == 200:
            logger.info(f"📧 Gmail API: correo enviado a {destinatarios} desde {from_email}")
            return True
        else:
            logger.error(f"Gmail API error {send_resp.status_code}: {send_resp.text}")
            return False

    except Exception as e:
        logger.error(f"Error en Gmail API: {e}", exc_info=True)
        return False


# ══════════════════════════════════════════════════════════════════════
# 📧 DISPATCHER PRINCIPAL — 3 prioridades en cascada
# ══════════════════════════════════════════════════════════════════════
def enviar_correo_con_smtp(destinatarios, asunto, html_mensaje, cuerpo_texto="",
                            imagen_final_buffer=None, evento=None):
    """
    Envía correo eligiendo automáticamente el mejor método disponible:
      1. SMTP personalizado del evento (si está configurado)
      2. Gmail API del organizador asignado (OAuth2)     ← nuevo
      3. Gmail del SUPERADMIN en settings.py (fallback)
    """
    if not isinstance(destinatarios, list):
        destinatarios = [destinatarios]

    SMTP_DEFAULTS = {'', 'apikey', None}

    # ─── PRIORIDAD 1: SMTP propio del evento ───────────────────────
    if evento and evento.smtp_password and evento.smtp_password not in SMTP_DEFAULTS:
        return _enviar_via_smtp(
            destinatarios, asunto, html_mensaje, cuerpo_texto, imagen_final_buffer,
            host=evento.smtp_host or settings.EMAIL_HOST,
            port=evento.smtp_port or settings.EMAIL_PORT,
            user=evento.smtp_user or settings.EMAIL_HOST_USER,
            password=evento.smtp_password,
            from_email=evento.default_from_email or settings.DEFAULT_FROM_EMAIL,
        )

    # ─── PRIORIDAD 2: Gmail API del organizador (OAuth2) ───────────
    if evento and settings.GOOGLE_CLIENT_ID:
        organizador = (
            PerfilUsuario.objects
            .filter(eventos=evento, rol='ORGANIZADOR')
            .exclude(google_refresh_token__isnull=True)
            .exclude(google_refresh_token='')
            .first()
        )
        if organizador:
            ok = enviar_con_gmail_api(
                destinatarios=destinatarios,
                asunto=asunto,
                html_mensaje=html_mensaje,
                imagen_final_buffer=imagen_final_buffer,
                from_email=organizador.google_email,
                refresh_token=organizador.google_refresh_token,
            )
            if ok:
                return True
            logger.warning("Gmail API del organizador falló. Usando Gmail del SUPERADMIN.")

    # ─── PRIORIDAD 3: Gmail SUPERADMIN (settings.py) ───────────────
    return _enviar_via_smtp(
        destinatarios, asunto, html_mensaje, cuerpo_texto, imagen_final_buffer,
        host=settings.EMAIL_HOST,
        port=settings.EMAIL_PORT,
        user=settings.EMAIL_HOST_USER,
        password=settings.EMAIL_HOST_PASSWORD,
        from_email=settings.DEFAULT_FROM_EMAIL,
    )


def enviar_correo_con_smtp_evento(participante, asunto, html_mensaje, imagen_final_buffer=None):
    if not participante.correo:
        logger.warning(f"Participante {participante.id} ({participante.nombres}) no tiene correo. Email no enviado.")
        return False
    return enviar_correo_con_smtp(
        destinatarios=[participante.correo],
        asunto=asunto,
        html_mensaje=html_mensaje,
        imagen_final_buffer=imagen_final_buffer,
        evento=participante.evento
    )


# ==========================================
# 🎟️ LISTADO DE PARTICIPANTES POR EVENTO
# ==========================================
@method_decorator(login_required(login_url='/participantes/login/'), name='dispatch')
class ParticipanteListView(ListView):
    model = Participante
    template_name = 'cliente/lista.html'
    context_object_name = 'Participante'
    paginate_by = 30
    ordering = ['-id']

    def dispatch(self, request, *args, **kwargs):
        evento_id = self.kwargs.get('evento_id')
        self.evento = get_object_or_404(Evento, pk=evento_id)
        
        perfil = get_object_or_404(PerfilUsuario, user=request.user)
        if perfil.rol != 'SUPERADMIN' and self.evento not in perfil.eventos.all():
            messages.error(request, "No tienes acceso a este evento.")
            return redirect('dashboard_eventos')
            
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = Participante.objects.filter(evento=self.evento).order_by('-id')
        q = self.request.GET.get("q")
        if q:
            queryset = queryset.filter(
                Q(nombres__icontains=q) | Q(dni__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evento'] = self.evento
        
        # 📈 Métricas de Ventas y Asistencia en Vivo
        total_vendidos = Participante.objects.filter(evento=self.evento, pago_confirmado=True).count()
        total_esperados = Participante.objects.filter(evento=self.evento).count()
        total_ingresados = Participante.objects.filter(evento=self.evento, entrada_usada=True).count()
        total_no_ingresados = max(0, total_vendidos - total_ingresados)
        ingresos_recaudados = Participante.objects.filter(evento=self.evento, pago_confirmado=True).aggregate(total=Sum('total_pagar'))['total'] or Decimal('0.00')

        # Conteo de categorías dinámicas para el gráfico de barras
        tarifas = Tarifa.objects.filter(evento=self.evento)
        categoria_datos = []
        for t in tarifas:
            count = Participante.objects.filter(evento=self.evento, tipo_entrada=t.tipo_entrada, pago_confirmado=True).count()
            categoria_datos.append({
                'label': t.tipo_entrada,
                'count': count
            })

        context['total_vendidos'] = total_vendidos
        context['total_esperados'] = total_esperados
        context['total_ingresados'] = total_ingresados
        context['total_no_ingresados'] = total_no_ingresados
        context['ingresos_recaudados'] = ingresos_recaudados
        context['categoria_datos'] = categoria_datos
        
        return context


# ==========================================
# 🎟️ CREAR / EDITAR / ELIMINAR PARTICIPANTES
# ==========================================
@method_decorator(login_required(login_url='/participantes/login/'), name='dispatch')
class ParticipanteCreateView(CreateView):
    model = Participante
    form_class = ParticipanteForm
    template_name = 'cliente/participante_form.html'

    def dispatch(self, request, *args, **kwargs):
        evento_id = self.kwargs.get('evento_id')
        self.evento = get_object_or_404(Evento, pk=evento_id)
        if not verificar_permiso_evento(request.user, self.evento):
            messages.error(request, "No tienes autorización para acceder a este evento.")
            return redirect('dashboard_eventos')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['evento'] = self.evento
        return kwargs

    def get_success_url(self):
        return reverse('participante_lista', kwargs={'evento_id': self.evento.pk})

    def form_valid(self, form):
        cantidad = form.cleaned_data.get('cantidad', 1) or 1
        total_actual = Participante.objects.filter(evento=self.evento).count()
        
        # 🛡️ Validar Aforo Máximo
        if self.evento.aforo_maximo and (total_actual + cantidad > self.evento.aforo_maximo):
            messages.error(self.request, f"❌ No se puede registrar al participante. Se supera el aforo máximo del evento ({self.evento.aforo_maximo} personas).")
            return self.form_invalid(form)

        # 🛡️ Validar Límite por Persona (Antireventa)
        if self.evento.limite_entradas_persona and (cantidad > self.evento.limite_entradas_persona):
            messages.error(self.request, f"❌ No se puede registrar. La cantidad de entradas ({cantidad}) supera el límite por persona ({self.evento.limite_entradas_persona}).")
            return self.form_invalid(form)

        participante = form.save(commit=False)
        participante.evento = self.evento

        # Asignar tarifa y precio
        tipo_entrada = form.cleaned_data.get('tipo_entrada')
        tarifa = Tarifa.objects.filter(evento=self.evento, tipo_entrada=tipo_entrada).first()
        if tarifa:
            participante.tarifa = tarifa

        precio_final = self.request.POST.get("precio_final")
        if precio_final:
            try:
                participante.precio = Decimal(precio_final)
            except Exception:
                participante.precio = Decimal("0.00")

        participante.save()

        # vouchers
        vouchers = self.request.FILES.getlist('vouchers')
        for v in vouchers:
            Voucher.objects.create(participante=participante, imagen=v)

        if participante.pago_confirmado:
            enviar_entrada_participante(participante)
            messages.success(self.request, f"Participante '{participante.nombres}' agregado y entrada enviada con éxito.")
        else:
            messages.success(self.request, f"Participante '{participante.nombres}' agregado con éxito (pago pendiente).")
            
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evento'] = self.evento
        return context


@method_decorator(login_required(login_url='/participantes/login/'), name='dispatch')
class ParticipanteUpdateView(UpdateView):
    model = Participante
    form_class = ParticipanteForm
    template_name = 'cliente/participante_form.html'

    def dispatch(self, request, *args, **kwargs):
        evento_id = self.kwargs.get('evento_id')
        self.evento = get_object_or_404(Evento, pk=evento_id)
        if not verificar_permiso_evento(request.user, self.evento):
            messages.error(request, "No tienes autorización para acceder a este evento.")
            return redirect('dashboard_eventos')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Participante.objects.filter(evento=self.evento)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['evento'] = self.evento
        return kwargs

    def get_success_url(self):
        return reverse('participante_lista', kwargs={'evento_id': self.evento.pk})

    def form_valid(self, form):
        participante = form.save(commit=False)

        tipo_entrada = form.cleaned_data.get('tipo_entrada')
        tarifa = Tarifa.objects.filter(evento=self.evento, tipo_entrada=tipo_entrada).first()
        if tarifa:
            participante.tarifa = tarifa

        precio_final = self.request.POST.get("precio_final")
        if precio_final:
            try:
                participante.precio = Decimal(precio_final)
            except Exception:
                participante.precio = Decimal("0.00")

        participante.save()

        vouchers = self.request.FILES.getlist('vouchers')
        for v in vouchers:
            Voucher.objects.create(participante=participante, imagen=v)

        if participante.pago_confirmado:
            enviar_entrada_participante(participante)
            messages.success(self.request, f"Participante '{participante.nombres}' actualizado y entrada enviada con éxito.")
        else:
            messages.success(self.request, f"Participante '{participante.nombres}' actualizado.")
            
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evento'] = self.evento
        return context


@method_decorator(login_required(login_url='/participantes/login/'), name='dispatch')
class ParticipanteDeleteView(DeleteView):
    model = Participante
    template_name = 'cliente/participante_confirm_delete.html'

    def dispatch(self, request, *args, **kwargs):
        evento_id = self.kwargs.get('evento_id')
        self.evento = get_object_or_404(Evento, pk=evento_id)
        if not verificar_permiso_evento(request.user, self.evento):
            messages.error(request, "No tienes autorización para acceder a este evento.")
            return redirect('dashboard_eventos')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Participante.objects.filter(evento=self.evento)

    def get_success_url(self):
        return reverse('participante_lista', kwargs={'evento_id': self.evento.pk})


# ==========================================
# 💵 CONFIRMAR PAGO & HELPER DE ENVÍO DE ENTRADAS
# ==========================================
def enviar_entrada_participante(participante):
    evento = participante.evento
    if not evento:
        return False

    # Generar QR dinámico
    base_url = settings.BASE_URL.rstrip("/")
    url_validacion = f"{base_url}/participantes/validar/{participante.token}/"
    qr_img = qrcode.make(url_validacion).convert("RGB")

    # Generar imagen combinada del boleto
    imagen_final = generar_imagen_personalizada(participante, qr_img)
    if not imagen_final:
        return False

    # Guardar localmente la entrada
    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
    buffer = BytesIO()
    imagen_final.save(buffer, format='PNG')
    buffer.seek(0)
    
    ruta_guardado = os.path.join(settings.MEDIA_ROOT, f"entrada_{participante.id}.png")
    imagen_final.save(ruta_guardado, format="PNG")

    # Enviar correo con SMTP dinámico
    asunto = f"🎟️ Tu entrada oficial para {evento.nombre}"
    html_mensaje = f"""
    <html><body>
        <p>Hola <strong>{participante.nombres}</strong>,</p>
        <p>Tu pago ha sido confirmado para <strong>{evento.nombre}</strong>.</p>
        <p>Cantidad de entradas adquiridas: <strong>{participante.cantidad}</strong>.</p>
        <p>Adjunto encontrarás tu boleto personalizado. Preséntalo impreso o en tu celular para ingresar.</p>
        <br>
        <img src="cid:entrada" style="max-width:100%; height:auto; border-radius: 12px;">
        <br>
        <p>¡Nos vemos pronto!</p>
    </body></html>
    """
    
    email_ok = enviar_correo_con_smtp_evento(participante, asunto, html_mensaje, buffer)

    # Subir imagen a ImgBB para envío por Twilio WhatsApp
    imgbb_url = None
    if evento.imgbb_api_key:
        try:
            encoded_image = base64.b64encode(buffer.getvalue()).decode("utf-8")
            resp = requests.post(
                "https://api.imgbb.com/1/upload", 
                data={"key": evento.imgbb_api_key, "image": encoded_image},
                timeout=15
            )
            if resp.status_code == 200:
                imgbb_url = resp.json()["data"]["url"]
        except Exception as e:
            logger.error(f"Error subiendo imagen a ImgBB: {e}")

    # 📱 Despachar WhatsApp según el proveedor configurado (Dinámico y Extensible)
    if participante.celular:
        provider = getattr(evento, 'whatsapp_provider', 'INACTIVE')
        
        num_limpio = "".join(filter(str.isdigit, participante.celular))
        if not num_limpio.startswith("51"):
            num_limpio = "51" + num_limpio

        msg_body = (
            f"¡Hola {participante.nombres}! 👋\n\n"
            f"Tu pago fue confirmado con éxito para *{evento.nombre}* ✅\n"
            f"Adquiriste {participante.cantidad} boletos 🎟️.\n\n"
            f"Adjuntamos tu entrada digital para el ingreso. ¡Te esperamos! 🚀"
        )

        if provider == 'TWILIO' and evento.twilio_account_sid and evento.twilio_auth_token:
            try:
                client = Client(evento.twilio_account_sid, evento.twilio_auth_token)
                wsp_num = f"whatsapp:{evento.twilio_whatsapp_number or evento.twilio_phone_number}"
                wsp_dest = f"whatsapp:+{num_limpio}"

                if imgbb_url:
                    client.messages.create(from_=wsp_num, to=wsp_dest, body=msg_body, media_url=[imgbb_url])
                else:
                    client.messages.create(from_=wsp_num, to=wsp_dest, body=msg_body)
                logger.info(f"📱 WhatsApp enviado vía Twilio a {num_limpio}")
            except Exception as e:
                logger.error(f"Error enviando mensaje WhatsApp Twilio: {e}")

        elif provider == 'CUSTOM_API' and evento.whatsapp_api_url:
            try:
                # 1. Armar headers
                headers = {"Content-Type": "application/json"}
                if evento.whatsapp_api_headers:
                    for line in evento.whatsapp_api_headers.splitlines():
                        if ":" in line:
                            k, v = line.split(":", 1)
                            headers[k.strip()] = v.strip()

                # 2. Armar y reemplazar variables en el payload
                payload_str = evento.whatsapp_api_payload or ""
                if not payload_str:
                    # Fallback simple
                    import json
                    payload_dict = {
                        "to": num_limpio,
                        "message": msg_body
                    }
                    if imgbb_url:
                        payload_dict["media_url"] = imgbb_url
                    payload_str = json.dumps(payload_dict)
                else:
                    payload_str = payload_str.replace("{celular}", num_limpio)
                    payload_str = payload_str.replace("{nombres}", participante.nombres or "")
                    payload_str = payload_str.replace("{evento}", evento.nombre)
                    payload_str = payload_str.replace("{entradas}", str(participante.cantidad))
                    payload_str = payload_str.replace("{url_imagen}", imgbb_url or "")

                # 3. Enviar petición HTTP POST
                try:
                    payload_json = json.loads(payload_str)
                    resp = requests.post(evento.whatsapp_api_url, json=payload_json, headers=headers, timeout=15)
                except ValueError:
                    resp = requests.post(evento.whatsapp_api_url, data=payload_str, headers=headers, timeout=15)

                logger.info(f"📱 WhatsApp Custom API enviado a {evento.whatsapp_api_url} - Status: {resp.status_code}")
            except Exception as e:
                logger.error(f"Error enviando WhatsApp Custom API: {e}")

    if email_ok:
        Participante.objects.filter(pk=participante.pk).update(email_enviado=True)

    return email_ok

@login_required(login_url='/participantes/login/')
def confirmar_pago(request, evento_id, pk):
    if request.method != 'POST':
        return redirect('participante_lista', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    participante = get_object_or_404(Participante, pk=pk, evento=evento)
    participante.pago_confirmado = True
    participante.save()

    enviar_entrada_participante(participante)
    messages.success(request, "✅ Pago confirmado y notificaciones (Correo / WhatsApp) despachadas.")
    return redirect('participante_lista', evento_id=evento.id)


# ==========================================
# 📧 ENVIAR MASIVO A TODOS LOS CONFIRMADOS
# ==========================================
@login_required(login_url='/participantes/login/')
def enviar_masivo(request, evento_id):
    if request.method != 'POST':
        return redirect('participante_lista', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    participantes = Participante.objects.filter(evento=evento, pago_confirmado=True)

    if not participantes.exists():
        messages.warning(request, "No hay participantes aprobados y con pago confirmado para enviar.")
        return redirect('participante_lista', evento_id=evento.id)

    enviados = 0
    for p in participantes:
        base_url = settings.BASE_URL.rstrip("/")
        url_val = f"{base_url}/participantes/validar/{p.token}/"
        qr_img = qrcode.make(url_val).convert("RGB")
        
        imagen_final = generar_imagen_personalizada(p, qr_img)
        if not imagen_final:
            continue
            
        buffer = BytesIO()
        imagen_final.save(buffer, format='PNG')
        buffer.seek(0)
        
        asunto = f"🎟️ Tu entrada para {evento.nombre}"
        html_mensaje = f"""
        <html><body>
            <p>Hola <strong>{p.nombres}</strong>,</p>
            <p>Aquí tienes tu entrada personalizada para <strong>{evento.nombre}</strong>.</p>
            <img src="cid:entrada" style="max-width:100%; height:auto;">
        </body></html>
        """
        
        if enviar_correo_con_smtp_evento(p, asunto, html_mensaje, buffer):
            enviados += 1

    messages.success(request, f"¡Despacho masivo finalizado con éxito! {enviados} correos enviados.")
    return redirect('participante_lista', evento_id=evento.id)


# ==========================================
# 📝 MEJORA: COMPONER IMAGEN DINÁMICA
# ==========================================
def generar_imagen_personalizada(participante, qr_img):
    evento = participante.evento
    
    # 1. Cargar imagen de fondo
    base_path = os.path.join(settings.BASE_DIR, 'cliente', 'static', 'img', 'asesor.jpeg')
    if evento and evento.imagen_fondo:
        base_path = evento.imagen_fondo.path
        
    if not os.path.exists(base_path):
        return None
        
    fondo = Image.open(base_path).convert("RGBA")

    # 2. Posicionar QR — usar config del evento o defaults
    pos_x     = evento.qr_pos_x if evento else 168
    pos_y     = evento.qr_pos_y if evento else 405
    qr_width  = evento.qr_ancho if evento else 567
    qr_height = evento.qr_alto  if evento else 569

    color_fondo = (evento.qr_color_fondo if evento else '#ffffff')
    transparente = (color_fondo == 'transparent')

    color_frente = (evento.qr_color_frente if evento else '#000000')

    # Re-colorear QR pixel a pixel: módulos → color_frente, fondo → color_fondo o transparente
    qr_rgba = qr_img.convert("RGBA").resize((qr_width, qr_height), Image.Resampling.LANCZOS)
    pixels = list(qr_rgba.getdata())
    try:
        cf = color_frente.lstrip('#')
        cf_rgb = (int(cf[0:2], 16), int(cf[2:4], 16), int(cf[4:6], 16))
    except Exception:
        cf_rgb = (0, 0, 0)
    if transparente:
        pixels = [(cf_rgb[0], cf_rgb[1], cf_rgb[2], 255) if not (r > 200 and g > 200 and b > 200)
                  else (255, 255, 255, 0) for r, g, b, a in pixels]
    else:
        try:
            cb = color_fondo.lstrip('#')
            cb_rgb = (int(cb[0:2], 16), int(cb[2:4], 16), int(cb[4:6], 16))
        except Exception:
            cb_rgb = (255, 255, 255)
        pixels = [(cf_rgb[0], cf_rgb[1], cf_rgb[2], 255) if not (r > 200 and g > 200 and b > 200)
                  else (cb_rgb[0], cb_rgb[1], cb_rgb[2], 255) for r, g, b, a in pixels]
    qr_rgba.putdata(pixels)
    entrada_completa = fondo.copy()
    capa = Image.new("RGBA", fondo.size, (0, 0, 0, 0))
    capa.paste(qr_rgba, (pos_x, pos_y), qr_rgba)
    entrada_completa = Image.alpha_composite(entrada_completa, capa)

    entrada_completa = entrada_completa.convert("RGB")
    
    # 3. Dibujar Nombre del Participante debajo del QR
    draw = ImageDraw.Draw(entrada_completa)
    nombre = (participante.nombres or "").upper()
    
    fuente_archivo = (evento.qr_fuente if evento and evento.qr_fuente else 'Roboto-Bold.ttf')
    font_path = os.path.join(settings.BASE_DIR, "cliente", "static", "fonts", fuente_archivo)
    if not os.path.exists(font_path):
        font_path = os.path.join(settings.BASE_DIR, "cliente", "static", "fonts", "Roboto-Bold.ttf")
        
    # Tamaño de fuente automático para que encaje
    font_size = 120
    while font_size > 40:
        try:
            font = ImageFont.truetype(font_path, font_size)
        except Exception:
            font = ImageFont.load_default()
            break

        bbox = draw.textbbox((0, 0), nombre, font=font)
        text_width = bbox[2] - bbox[0]
        if text_width <= (qr_width - 30):
            break
        font_size -= 8

    # Centrar y dibujar texto
    try:
        font = ImageFont.truetype(font_path, font_size)
    except Exception:
        font = ImageFont.load_default()
        
    bbox = draw.textbbox((0, 0), nombre, font=font)
    text_width = bbox[2] - bbox[0]
    
    texto_x = pos_x + (qr_width // 2) - (text_width // 2)
    texto_y = pos_y + qr_height + 40
    
    draw.text(
        (texto_x, texto_y),
        nombre,
        font=font,
        fill="white",
        stroke_width=6,
        stroke_fill="black"
    )
    
    return entrada_completa


# ==========================================
# 📥 EXCEL IMPORT DE ACUERDO CON TARIFAS
# ==========================================
@login_required(login_url='/participantes/login/')
def importar_excel(request, evento_id):
    evento = get_object_or_404(Evento, pk=evento_id)
    if request.method == "POST" and request.FILES.get('excel_file'):
        import pandas as pd
        archivo = request.FILES['excel_file']
        try:
            df = pd.read_excel(archivo)
            df.columns = df.columns.str.strip()
            
            # Normalizar columnas
            df = df.rename(columns={
                'Nombre': 'Nombre',
                'DNI': 'DNI', 
                'TELEFONO': 'TELEFONO',
                'Correo electrónico': 'Correo',
                'ASESOR QUE TE INVITO': 'Vendedor',
                'Tipo de entrada': 'Tipo_Entrada'
            })

            # Cargar Tarifas dinámicas del Evento
            tarifas = Tarifa.objects.filter(evento=evento)
            tarifas_dict = {}
            for t in tarifas:
                tarifas_dict[t.tipo_entrada.upper()] = {
                    "PREVENTA1": t.preventa_1,
                    "PREVENTA2": t.preventa_2,
                    "PREVENTA3": t.preventa_3,
                    "PUERTA": t.puerta
                }

            first_tariff_name = tarifas.first().tipo_entrada.upper() if tarifas.exists() else "EMPRENDEDOR"
            enviados = 0
            errores = 0
            
            for _, row in df.iterrows():
                try:
                    if pd.isna(row.get('DNI')) or pd.isna(row.get('Nombre')):
                        continue
                    
                    telefono = ''
                    if not pd.isna(row.get('TELEFONO')):
                        telefono = str(row.get('TELEFONO')).replace('.0', '').strip()
                    
                    tipo_texto = str(row.get('Tipo_Entrada', '')).strip().upper()
                    if not tipo_texto:
                        tipo_texto = first_tariff_name
                    tipo_texto = tipo_texto.replace("ACCES", "ACCESS")
                    
                    # Buscar coincidencia
                    tipo_encontrado = first_tariff_name
                    for key in tarifas_dict.keys():
                        if key in tipo_texto:
                            tipo_encontrado = key
                            break
                            
                    # Buscar preventa
                    preventa_tipo = "PREVENTA1"
                    for pv in ["PREVENTA1", "PREVENTA2", "PREVENTA3", "PUERTA"]:
                        if pv in tipo_texto:
                            preventa_tipo = pv
                            break
                            
                    precio = Decimal("0.00")
                    if tipo_encontrado in tarifas_dict:
                        precio = tarifas_dict[tipo_encontrado].get(preventa_tipo, Decimal("0.00"))
                    
                    tarifa_db = Tarifa.objects.filter(evento=evento, tipo_entrada=tipo_encontrado).first()
                    
                    part = Participante.objects.create(
                        evento=evento,
                        tarifa=tarifa_db,
                        nombres=row['Nombre'],
                        apellidos="",
                        dni=str(row['DNI']),
                        celular=telefono,
                        correo=row['Correo'] if not pd.isna(row.get('Correo')) else '',
                        vendedor=row['Vendedor'] if not pd.isna(row.get('Vendedor')) else '',
                        tipo_entrada=tipo_encontrado,
                        cantidad=1,
                        precio=precio,
                        pago_confirmado=True
                    )
                    enviados += 1
                except Exception as e:
                    errores += 1
                    logger.error(f"Fallo importando participante: {e}")
            
            messages.success(request, f"✅ {enviados} participantes importados. Usa 'Enviar tickets pendientes' para despachar los boletos por correo.")
        except Exception as e:
            messages.error(request, f"❌ Error de lectura de Excel: {e}")
            
    return redirect('participante_lista', evento_id=evento.id)


# ==========================================
# 📧 ENVIAR SIGUIENTE TICKET PENDIENTE (AJAX one-by-one)
# ==========================================
@login_required(login_url='/participantes/login/')
def enviar_siguiente_pendiente(request, evento_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    evento = get_object_or_404(Evento, pk=evento_id)
    pendiente = Participante.objects.filter(
        evento=evento, pago_confirmado=True, email_enviado=False
    ).first()

    if not pendiente:
        total = Participante.objects.filter(evento=evento, pago_confirmado=True).count()
        return JsonResponse({'status': 'done', 'pendientes': 0, 'total': total})

    ok = enviar_entrada_participante(pendiente)

    pendientes = Participante.objects.filter(evento=evento, pago_confirmado=True, email_enviado=False).count()
    total = Participante.objects.filter(evento=evento, pago_confirmado=True).count()
    return JsonResponse({
        'status': 'ok' if ok else 'error',
        'nombre': pendiente.nombres,
        'pendientes': pendientes,
        'total': total,
    })


# ==========================================
# 📤 EXPORTAR EXCEL POR EVENTO
# ==========================================
@login_required(login_url='/participantes/login/')
def exportar_excel(request, evento_id):
    import pandas as pd
    evento = get_object_or_404(Evento, pk=evento_id)
    participantes = Participante.objects.filter(evento=evento).values().order_by('id')

    if not participantes:
        return HttpResponse("No hay datos que exportar.", content_type="text/plain")

    df = pd.DataFrame(participantes)
    if 'paquete' in df.columns:
        df = df.drop(columns=['paquete'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Participantes')
        workbook = writer.book
        worksheet = writer.sheets['Participantes']
        
        # Formatear
        header_fill = PatternFill(start_color="1A0033", end_color="1A0033", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            
    response = HttpResponse(
        output.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Participantes_{evento.nombre.replace(" ", "_")}.xlsx'
    return response


# ==========================================
# 🧹 LIMPIAR HISTORIAL DE ENTRADAS
# ==========================================
@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def limpiar_historial(request, evento_id):
    evento = get_object_or_404(Evento, pk=evento_id)
    if request.method == "POST":
        participantes = Participante.objects.filter(evento=evento)
        for p in participantes:
            if p.qr:
                try:
                    if os.path.exists(p.qr.path):
                        os.remove(p.qr.path)
                except Exception as e:
                    logger.error(f"Error borrando archivo QR: {e}")
            for v in p.vouchers.all():
                if v.imagen:
                    try:
                        if os.path.exists(v.imagen.path):
                            os.remove(v.imagen.path)
                    except Exception as e:
                        logger.error(f"Error borrando archivo voucher: {e}")

        # Limpiar imágenes combinadas locales
        try:
            for filename in os.listdir(settings.MEDIA_ROOT):
                if filename.startswith("entrada_") and filename.endswith(".png"):
                    try:
                        os.remove(os.path.join(settings.MEDIA_ROOT, filename))
                    except Exception as e:
                        logger.error(f"Error borrando archivo local: {e}")
        except Exception as e:
            logger.error(f"Error limpiando media root local: {e}")

        participantes.delete()
        messages.success(request, "🧹 El historial de entradas de este evento ha sido eliminado por completo.")
        
    return redirect('participante_lista', evento_id=evento.id)


# ==========================================
# 👥 PRE-REGISTRO & PREVIA PARTICIPANTES
# ==========================================
@login_required(login_url='/participantes/login/')
def registro_participante(request, evento_id):
    evento = get_object_or_404(Evento, pk=evento_id)

    ultimo = Previaparticipantes.objects.filter(evento=evento).annotate(
        cod_num=Cast(Substr('cod_part', 4), IntegerField())
    ).aggregate(max_cod=Max('cod_num'))

    siguiente_numero = (ultimo['max_cod'] or 0) + 1
    nuevo_cod = f"CLI{siguiente_numero:03d}"

    queryset = Previaparticipantes.objects.filter(evento=evento).annotate(
        cod_num=Cast(Substr('cod_part', 4), IntegerField())
    ).order_by('-cod_num')

    q = request.GET.get('q')
    if q:
        queryset = queryset.filter(
            Q(nombres__icontains=q) | Q(dni__icontains=q) | Q(celular__icontains=q)
        )

    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            contador = siguiente_numero
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 4 or not row[0]:   # saltar filas vacías o sin nombre
                    continue
                nombres, dni, celular, correo = row[:4]
                Previaparticipantes.objects.create(
                    evento=evento,
                    cod_part=f"CLI{contador:03d}",
                    nombres=nombres,
                    dni=dni,
                    celular=celular,
                    correo=correo
                )
                contador += 1
            messages.success(request, "Participantes importados al registro de la previa.")
        else:
            Previaparticipantes.objects.create(
                evento=evento,
                cod_part=nuevo_cod,
                nombres=request.POST.get('nombres'),
                dni=request.POST.get('dni'),
                celular=request.POST.get('celular'),
                correo=request.POST.get('correo')
            )
            messages.success(request, "Participante registrado en el pre-registro.")
        return redirect('registro_participante', evento_id=evento.id)

    return render(request, 'cliente/registro_participante.html', {
        'evento': evento,
        'nuevo_cod': nuevo_cod,
        'page_obj': page_obj
    })


@login_required(login_url='/participantes/login/')
def actualizar_participante_previa(request, evento_id, pk):
    evento = get_object_or_404(Evento, pk=evento_id)
    participante = get_object_or_404(Previaparticipantes, pk=pk, evento=evento)
    if request.method == "POST":
        participante.nombres = request.POST.get("nombres")
        participante.dni = request.POST.get("dni")
        participante.celular = request.POST.get("celular")
        participante.correo = request.POST.get("correo")
        participante.save()
        messages.success(request, "Datos de pre-registro actualizados.")
        return redirect('registro_participante', evento_id=evento.id)
    return render(request, 'cliente/actualizar_participante_previo.html', {
        'participante': participante,
        'evento': evento
    })


@login_required(login_url='/participantes/login/')
def eliminar_participante_previa(request, evento_id, pk):
    evento = get_object_or_404(Evento, pk=evento_id)
    participante = get_object_or_404(Previaparticipantes, pk=pk, evento=evento)
    if request.method == "POST":
        if participante.qr_image:
            try:
                if os.path.exists(participante.qr_image.path):
                    os.remove(participante.qr_image.path)
            except Exception as e:
                logger.error(f"Error borrando archivo QR previa: {e}")
        participante.delete()
        messages.success(request, "Registro eliminado.")
    return redirect('registro_participante', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def limpiar_historial_previa(request, evento_id):
    evento = get_object_or_404(Evento, pk=evento_id)
    if request.method == "POST":
        previa_parts = Previaparticipantes.objects.filter(evento=evento)
        for p in previa_parts:
            if p.qr_image:
                try:
                    if os.path.exists(p.qr_image.path):
                        os.remove(p.qr_image.path)
                except Exception as e:
                    logger.error(f"Error borrando archivo QR previa: {e}")
        previa_parts.delete()
        messages.success(request, "🧹 Historial de pre-registro limpiado con éxito.")
    return redirect('registro_participante', evento_id=evento.id)


# ==========================================
# 🔒 VALIDACIÓN QR INTEGRADA (GOOGLE CÁMARA)
# ==========================================
@login_required(login_url='/participantes/login/')
def validar_entrada(request, token):
    participante = get_object_or_404(Participante, token=token)
    evento = participante.evento

    perfil = get_object_or_404(PerfilUsuario, user=request.user)
    if perfil.rol != 'SUPERADMIN' and evento not in perfil.eventos.all():
        messages.error(request, "No tienes permisos de validación para este evento.")
        return redirect('dashboard_eventos')

    valido = False
    if not participante.entrada_usada:
        participante.entrada_usada = True
        participante.hora_ingreso = timezone.now()
        participante.save()
        valido = True
        mensaje = "✅ ¡Acceso Autorizado! Bienvenido al evento."
    else:
        if participante.hora_ingreso:
            mensaje = f"❌ ¡Boleto ya Utilizado! Registrado el {participante.hora_ingreso.strftime('%d/%m/%Y %I:%M %p')}"
        else:
            mensaje = "❌ ¡Boleto ya Utilizado! (sin fecha de ingreso registrada)"

    return render(request, 'cliente/entrada_valida.html' if valido else 'cliente/entrada_usada.html', {
        'participante': participante,
        'evento': evento,
        'mensaje': mensaje,
        'fecha_ingreso': participante.hora_ingreso
    })


@login_required(login_url='/participantes/login/')
def validar_entrada_previo(request, token):
    participante = get_object_or_404(Previaparticipantes, token=token)
    evento = participante.evento

    perfil = get_object_or_404(PerfilUsuario, user=request.user)
    if perfil.rol != 'SUPERADMIN' and evento not in perfil.eventos.all():
        messages.error(request, "No tienes permisos de validación para este evento.")
        return redirect('dashboard_eventos')

    valido = False
    if not participante.entrada_usada:
        participante.entrada_usada = True
        participante.hora_ingreso = timezone.now()
        participante.save()
        valido = True
        mensaje = "✅ ¡Acceso Previa Autorizado!"
    else:
        if participante.hora_ingreso:
            mensaje = f"❌ ¡Boleto de Previa ya Utilizado! Registrado el {participante.hora_ingreso.strftime('%d/%m/%Y %I:%M %p')}"
        else:
            mensaje = "❌ ¡Boleto de Previa ya Utilizado! (sin fecha de ingreso registrada)"

    return render(request, 'cliente/entrada_valida.html' if valido else 'cliente/entrada_usada.html', {
        'participante': participante,
        'evento': evento,
        'mensaje': mensaje,
        'fecha_ingreso': participante.hora_ingreso
    })


# ==========================================
# 🔗 VISTAS SECUNDARIAS COMPATIBLES
# ==========================================
@login_required(login_url='/participantes/login/')
def check_admin_masivo(request, evento_id):
    if request.method != 'POST':
        return redirect('participante_lista', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    Participante.objects.filter(evento=evento).update(validado_admin=True)
    messages.success(request, "Administración validada para todas las entradas de este evento.")
    return redirect('participante_lista', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def check_contabilidad_masivo(request, evento_id):
    if request.method != 'POST':
        return redirect('participante_lista', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    Participante.objects.filter(evento=evento).update(validado_contabilidad=True)
    messages.success(request, "Contabilidad validada para todas las entradas de este evento.")
    return redirect('participante_lista', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def marcar_ingreso(request, evento_id, pk):
    """Marca ingreso manual para un Participante (lista principal)."""
    if request.method != 'POST':
        return redirect('participante_lista', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    participante = get_object_or_404(Participante, pk=pk, evento=evento)
    if not participante.entrada_usada:
        participante.entrada_usada = True
        participante.hora_ingreso = timezone.now()
        participante.save()
        messages.success(request, f"Entrada marcada como ingresada para {participante.nombres}.")
    return redirect('participante_lista', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def convertir_previa_a_participante(request, evento_id, pk):
    """
    Convierte un Previaparticipante en un Participante confirmado.
    Genera el boleto y lo envía por correo automáticamente.
    El pre-registro queda marcado como 'enviado' para indicar que fue procesado.
    """
    evento    = get_object_or_404(Evento, pk=evento_id)
    previa    = get_object_or_404(Previaparticipantes, pk=pk, evento=evento)

    if request.method == 'POST':
        tipo_entrada = request.POST.get('tipo_entrada', '').strip()
        precio_str   = request.POST.get('precio_final', '0')

        try:
            precio = Decimal(precio_str)
        except Exception:
            precio = Decimal('0.00')

        # Verificar que no exista ya un participante con el mismo DNI en este evento
        if previa.dni and Participante.objects.filter(evento=evento, dni=previa.dni).exists():
            messages.warning(
                request,
                f"'{previa.nombres}' (DNI: {previa.dni}) ya existe en la lista de participantes de este evento."
            )
            return redirect('registro_participante', evento_id=evento.id)

        tarifa = Tarifa.objects.filter(evento=evento, tipo_entrada=tipo_entrada).first()

        # Crear el Participante con los datos del pre-registro
        nuevo = Participante(
            evento       = evento,
            tarifa       = tarifa,
            nombres      = previa.nombres or '',
            apellidos    = '',
            dni          = previa.dni or '',
            celular      = previa.celular or '',
            correo       = previa.correo or '',
            tipo_entrada = tipo_entrada,
            cantidad     = 1,
            precio       = precio,
            pago_confirmado          = True,
            validado_admin           = False,
            validado_contabilidad    = False,
        )
        nuevo.save()   # genera cod_cliente, token y QR automáticamente

        # Enviar boleto por correo / WhatsApp
        enviar_entrada_participante(nuevo)

        # Marcar el pre-registro como procesado (sin borrarlo)
        previa.enviado = True
        previa.save()

        messages.success(
            request,
            f"✅ '{previa.nombres}' registrado como participante confirmado y boleto enviado a {previa.correo}."
        )
        return redirect('registro_participante', evento_id=evento.id)

    # GET → mostrar modal inline (no se usa directo, el modal está en el template)
    return redirect('registro_participante', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def marcar_ingreso_previa(request, evento_id, pk):
    """Marca ingreso manual para un Previaparticipante (pre-registro)."""
    if request.method != 'POST':
        return redirect('registro_participante', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    participante = get_object_or_404(Previaparticipantes, pk=pk, evento=evento)
    if not participante.entrada_usada:
        participante.entrada_usada = True
        participante.hora_ingreso = timezone.now()
        participante.save()
        messages.success(request, f"Ingreso marcado para {participante.nombres}.")
    return redirect('registro_participante', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def reenviar_correo(request, evento_id, pk):
    if request.method != 'POST':
        return redirect('participante_lista', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    participante = get_object_or_404(Participante, pk=pk, evento=evento)
    
    base_url = settings.BASE_URL.rstrip("/")
    url_val = f"{base_url}/participantes/validar/{participante.token}/"
    qr_img = qrcode.make(url_val).convert("RGB")
    
    imagen_final = generar_imagen_personalizada(participante, qr_img)
    if not imagen_final:
        messages.error(request, "No se pudo generar la imagen del boleto.")
        return redirect('participante_lista', evento_id=evento.id)
        
    buffer = BytesIO()
    imagen_final.save(buffer, format='PNG')
    buffer.seek(0)
    
    asunto = f"🎟️ Reenvío de tu entrada oficial para {evento.nombre}"
    html_mensaje = f"""
    <html><body>
        <p>Hola <strong>{participante.nombres}</strong>,</p>
        <p>Este es un reenvío de tu entrada oficial para <strong>{evento.nombre}</strong>.</p>
        <img src="cid:entrada" style="max-width:100%; height:auto;">
    </body></html>
    """
    
    if enviar_correo_con_smtp_evento(participante, asunto, html_mensaje, buffer):
        messages.success(request, "Boleto reenviado por correo con éxito.")
    else:
        messages.error(request, "Fallo al enviar el correo.")
        
    return redirect('participante_lista', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def enviar_whatsapp_qr(request, evento_id, cod_part):
    if request.method != 'POST':
        return redirect('registro_participante', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    p = get_object_or_404(Previaparticipantes, cod_part=cod_part, evento=evento)
    p.enviado = True
    p.save()
    messages.success(request, f"QR marcado como enviado para {p.nombres}.")
    return redirect('registro_participante', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def enviar_todos_whatsapp(request, evento_id):
    if request.method != 'POST':
        return redirect('registro_participante', evento_id=evento_id)
    evento = get_object_or_404(Evento, pk=evento_id)
    previa_parts = Previaparticipantes.objects.filter(evento=evento, enviado=False)
    for p in previa_parts:
        p.enviado = True
        p.save()
    messages.success(request, "Envío masivo procesado para este evento.")
    return redirect('registro_participante', evento_id=evento.id)


@login_required(login_url='/participantes/login/')
def exportar_excel_previo(request, evento_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    evento = get_object_or_404(Evento, pk=evento_id)
    participantes = Previaparticipantes.objects.filter(evento=evento).order_by('cod_part')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-Registro"
    
    # Header row
    headers = ["Código", "Nombres", "DNI", "Celular", "Correo", "Enviado", "Ingresado"]
    header_fill = PatternFill(start_color="1A0033", end_color="1A0033", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_num, p in enumerate(participantes, 2):
        ws.cell(row=row_num, column=1, value=p.cod_part)
        ws.cell(row=row_num, column=2, value=p.nombres)
        ws.cell(row=row_num, column=3, value=p.dni)
        ws.cell(row=row_num, column=4, value=p.celular)
        ws.cell(row=row_num, column=5, value=p.correo)
        ws.cell(row=row_num, column=6, value="Sí" if p.enviado else "No")
        ws.cell(row=row_num, column=7, value="Sí" if p.entrada_usada else "No")
    
    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 3)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=PreRegistro_{evento.nombre.replace(" ", "_")}.xlsx'
    return response


@login_required(login_url='/participantes/login/')
def exportar_pdf_previo(request, evento_id):
    evento = get_object_or_404(Evento, pk=evento_id)
    participantes = Previaparticipantes.objects.filter(evento=evento).order_by('cod_part')
    
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=16, spaceAfter=12, textColor=colors.HexColor('#7b1fa2'))
        elements.append(Paragraph(f"Pre-Registro — {evento.nombre}", title_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Table header
        data = [["Código", "Nombres", "DNI", "Celular", "Correo", "Enviado", "Ingresado"]]
        for p in participantes:
            data.append([
                p.cod_part,
                (p.nombres or "")[:40],
                p.dni or "",
                p.celular or "",
                (p.correo or "")[:35],
                "Sí" if p.enviado else "No",
                "Sí" if p.entrada_usada else "No",
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A0033')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9F0FF'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWHEIGHT', (0, 0), (-1, -1), 18),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=PreRegistro_{evento.nombre.replace(" ", "_")}.pdf'
        return response
    
    except ImportError:
        # Fallback: generate HTML table for browser printing
        participantes_data = list(participantes.values('cod_part', 'nombres', 'dni', 'celular', 'correo', 'enviado', 'entrada_usada'))
        filas_html = ""
        for p in participantes_data:
            filas_html += f"""<tr>
                <td>{p['cod_part']}</td><td>{p['nombres'] or ''}</td>
                <td>{p['dni'] or ''}</td><td>{p['celular'] or ''}</td>
                <td>{p['correo'] or ''}</td>
                <td>{'Sí' if p['enviado'] else 'No'}</td>
                <td>{'Sí' if p['entrada_usada'] else 'No'}</td>
            </tr>"""
        
        html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
        <title>Pre-Registro {evento.nombre}</title>
        <style>body{{font-family:Arial;font-size:12px;}}
        table{{width:100%;border-collapse:collapse;}}th{{background:#1A0033;color:white;padding:8px;}}
        td{{padding:6px;border:1px solid #ccc;}}tr:nth-child(even){{background:#f9f0ff;}}
        </style></head><body>
        <h2>Pre-Registro — {evento.nombre}</h2>
        <table><thead><tr><th>Código</th><th>Nombres</th><th>DNI</th><th>Celular</th><th>Correo</th><th>Enviado</th><th>Ingresado</th></tr></thead>
        <tbody>{filas_html}</tbody></table>
        <script>window.print();</script></body></html>"""
        return HttpResponse(html, content_type='text/html')


# ==========================================
# 👥 GESTIÓN DE USUARIOS (SUPERADMIN ONLY)
# ==========================================
from django.contrib.auth.models import User

@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def usuario_lista(request):
    perfil = get_object_or_404(PerfilUsuario, user=request.user)
    if perfil.rol == 'SUPERADMIN':
        usuarios = PerfilUsuario.objects.all().select_related('user').prefetch_related('eventos')
    else:
        mis_eventos = perfil.eventos.all()
        usuarios = PerfilUsuario.objects.filter(
            rol='REGISTRADOR',
            eventos__in=mis_eventos
        ).distinct().select_related('user').prefetch_related('eventos')
    return render(request, 'cliente/usuario_lista.html', {
        'usuarios': usuarios,
        'perfil': perfil
    })

@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def usuario_crear(request):
    perfil_admin = get_object_or_404(PerfilUsuario, user=request.user)
    
    if perfil_admin.rol == 'SUPERADMIN':
        eventos = Evento.objects.all()
    else:
        eventos = perfil_admin.eventos.all()
        
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        
        if perfil_admin.rol == 'SUPERADMIN':
            rol = request.POST.get("rol", "REGISTRADOR")
            eventos_ids = request.POST.getlist("eventos")
        else:
            rol = "REGISTRADOR"
            eventos_ids = request.POST.getlist("eventos")
            mis_eventos_ids = [str(e.id) for e in perfil_admin.eventos.all()]
            eventos_ids = [eid for eid in eventos_ids if eid in mis_eventos_ids]
            if not eventos_ids:
                messages.error(request, "Debes seleccionar al menos uno de tus eventos asignados.")
                return redirect('usuario_crear')
                
        if not username or not password or not email:
            messages.error(request, "Todos los campos (usuario, email y contraseña) son obligatorios.")
            return redirect('usuario_crear')
            
        if User.objects.filter(username=username).exists():
            messages.error(request, "El nombre de usuario ya existe.")
            return redirect('usuario_crear')
            
        user = User.objects.create_user(username=username, email=email, password=password)
        perfil = PerfilUsuario.objects.create(user=user, rol=rol)
        
        for ev_id in eventos_ids:
            evento = Evento.objects.filter(pk=ev_id).first()
            if evento:
                perfil.eventos.add(evento)
                
        primer_evento = perfil.eventos.first()
        try:
            asunto = "🔐 Tus credenciales de acceso - Sistema de Entradas"
            eventos_lista_str = ", ".join([e.nombre for e in perfil.eventos.all()]) or "Ninguno"
            mensaje_html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.5;">
                <h2 style="color: #0ea5e9;">¡Hola {username}!</h2>
                <p>Se ha creado una cuenta para ti en la plataforma de <strong>Sistema de Entradas</strong>.</p>
                <p>Aquí tienes tus credenciales de acceso:</p>
                <table style="border-collapse: collapse; width: 100%; max-width: 400px; margin: 20px 0;">
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Usuario:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{username}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Contraseña:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{password}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Rol:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{perfil.get_rol_display()}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Eventos:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{eventos_lista_str}</td>
                    </tr>
                </table>
                <p>Puedes iniciar sesión en: <a href="{settings.BASE_URL}/login/">{settings.BASE_URL}/login/</a></p>
                <br>
                <p>Saludos,<br><strong>Soporte Técnico</strong></p>
            </body>
            </html>
            """
            
            destinatarios = [email, "eldespertardelemprendedor999@gmail.com"]
            cuerpo_txt = f"Hola {username}, tus credenciales son:\nUsuario: {username}\nContraseña: {password}\nRol: {perfil.get_rol_display()}"
            enviado = enviar_correo_con_smtp(
                destinatarios=destinatarios,
                asunto=asunto,
                html_mensaje=mensaje_html,
                cuerpo_texto=cuerpo_txt,
                evento=primer_evento
            )
            if enviado:
                messages.success(request, f"¡Usuario '{username}' creado y credenciales enviadas por correo con éxito!")
            else:
                messages.warning(request, f"Usuario '{username}' creado correctamente. No se pudo enviar el correo de credenciales (el organizador debe conectar su Gmail en 'Mi Cuenta').")
        except Exception as e:
            messages.warning(request, f"Usuario '{username}' creado pero ocurrió un error al enviar el correo: {e}")
            
        return redirect('usuario_lista')
        
    return render(request, 'cliente/usuario_form.html', {
        'perfil': perfil_admin,
        'eventos': eventos
    })

@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def usuario_editar(request, pk):
    perfil_admin = get_object_or_404(PerfilUsuario, user=request.user)
    perfil = get_object_or_404(PerfilUsuario, pk=pk)
    user = perfil.user
    
    if perfil_admin.rol == 'SUPERADMIN':
        eventos = Evento.objects.all()
    else:
        if perfil.rol != 'REGISTRADOR':
            messages.error(request, "No tienes permiso para editar este tipo de usuario.")
            return redirect('usuario_lista')
        mis_eventos = perfil_admin.eventos.all()
        shared_events = perfil.eventos.filter(id__in=mis_eventos.values_list('id', flat=True))
        if not shared_events.exists():
            messages.error(request, "No tienes acceso para editar este usuario.")
            return redirect('usuario_lista')
        eventos = mis_eventos
        
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "").strip()
        
        if perfil_admin.rol == 'SUPERADMIN':
            rol = request.POST.get("rol", perfil.rol)
            eventos_ids = request.POST.getlist("eventos")
        else:
            rol = "REGISTRADOR"
            eventos_ids = request.POST.getlist("eventos")
            mis_eventos_ids = [str(e.id) for e in perfil_admin.eventos.all()]
            eventos_ids = [eid for eid in eventos_ids if eid in mis_eventos_ids]
            if not eventos_ids:
                messages.error(request, "Debes seleccionar al menos uno de tus eventos asignados.")
                return redirect('usuario_editar', pk=pk)
                
        user.email = email
        if password:
            user.set_password(password)
        user.save()
        
        perfil.rol = rol
        perfil.eventos.clear()
        for ev_id in eventos_ids:
            evento = Evento.objects.filter(pk=ev_id).first()
            if evento:
                perfil.eventos.add(evento)
        perfil.save()
        
        primer_evento = perfil.eventos.first()
        try:
            asunto = "🔐 Actualización de tu cuenta - Sistema de Entradas"
            eventos_lista_str = ", ".join([e.nombre for e in perfil.eventos.all()]) or "Ninguno"
            pass_str = password if password else "(Sin cambios)"
            mensaje_html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.5;">
                <h2 style="color: #0ea5e9;">¡Hola {user.username}!</h2>
                <p>Se ha actualizado tu cuenta en la plataforma de <strong>Sistema de Entradas</strong>.</p>
                <p>Tus credenciales actualizadas son:</p>
                <table style="border-collapse: collapse; width: 100%; max-width: 400px; margin: 20px 0;">
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Usuario:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{user.username}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Contraseña:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{pass_str}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Rol:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{perfil.get_rol_display()}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; background: #f8fafc;">Eventos:</td>
                        <td style="padding: 8px; border: 1px solid #ddd;">{eventos_lista_str}</td>
                    </tr>
                </table>
                <br>
                <p>Saludos,<br><strong>Soporte Técnico</strong></p>
            </body>
            </html>
            """
            destinatarios = [email, "eldespertardelemprendedor999@gmail.com"]
            cuerpo_txt = f"Hola {user.username}, tu cuenta ha sido actualizada."
            enviado = enviar_correo_con_smtp(
                destinatarios=destinatarios,
                asunto=asunto,
                html_mensaje=mensaje_html,
                cuerpo_texto=cuerpo_txt,
                evento=primer_evento
            )
            if enviado:
                messages.success(request, f"¡Usuario '{user.username}' actualizado y credenciales enviadas por correo con éxito!")
            else:
                messages.warning(request, f"Usuario '{user.username}' actualizado correctamente. No se pudo enviar el correo de credenciales (el organizador debe conectar su Gmail en 'Mi Cuenta').")
        except Exception as e:
            messages.warning(request, f"Usuario '{user.username}' actualizado pero no se pudo enviar el correo: {e}")
            
        return redirect('usuario_lista')
        
    return render(request, 'cliente/usuario_form.html', {
        'perfil': perfil_admin,
        'perfil_edit': perfil,
        'eventos': eventos
    })

# ══════════════════════════════════════════════════════════════════════
# 🔐 GOOGLE OAUTH2 — Conectar Gmail por Organizador
# ══════════════════════════════════════════════════════════════════════

@login_required(login_url='/participantes/login/')
def mi_cuenta(request):
    """Página de perfil del usuario: muestra estado Gmail y permite conectar/desconectar."""
    perfil = get_object_or_404(PerfilUsuario, user=request.user)
    google_configurado = bool(getattr(settings, 'GOOGLE_CLIENT_ID', ''))
    return render(request, 'cliente/mi_cuenta.html', {
        'perfil': perfil,
        'google_configurado': google_configurado,
    })


# ══════════════════════════════════════════════════════════════════════
# 🎨 EDITOR VISUAL DE ENTRADA (QR Position + Color + Font)
# ══════════════════════════════════════════════════════════════════════
import json as _json

FUENTES_DISPONIBLES = [
    ('Roboto-Bold.ttf',    'Roboto Bold'),
    ('Poppins-Bold.ttf',   'Poppins Bold'),
    ('Poppins-Regular.ttf','Poppins Regular'),
    ('arial.ttf',          'Arial'),
    ('ARIALBD.TTF',        'Arial Bold'),
    ('ariblk.ttf',         'Arial Black'),
    ('ARIALN.TTF',         'Arial Narrow'),
    ('ARIALNB.TTF',        'Arial Narrow Bold'),
]


@login_required(login_url='/participantes/login/')
def editor_entrada(request, evento_id):
    evento = get_object_or_404(Evento, pk=evento_id)

    img_ancho, img_alto = 900, 1500
    base_path = os.path.join(settings.BASE_DIR, 'cliente', 'static', 'img', 'asesor.jpeg')
    if evento.imagen_fondo:
        base_path = evento.imagen_fondo.path
    if os.path.exists(base_path):
        with Image.open(base_path) as img:
            img_ancho, img_alto = img.size

    fondo_url = evento.imagen_fondo.url if evento.imagen_fondo else None

    return render(request, 'cliente/editor_entrada.html', {
        'evento': evento,
        'img_ancho': img_ancho,
        'img_alto': img_alto,
        'fondo_url': fondo_url,
        'fuentes': FUENTES_DISPONIBLES,
    })


@login_required(login_url='/participantes/login/')
def preview_entrada_ajax(request, evento_id):
    """Genera y devuelve la imagen del boleto con los params actuales (base64 JPEG)."""
    evento = get_object_or_404(Evento, pk=evento_id)

    try:
        pos_x        = int(request.GET.get('pos_x',         evento.qr_pos_x))
        pos_y        = int(request.GET.get('pos_y',         evento.qr_pos_y))
        ancho        = int(request.GET.get('ancho',         evento.qr_ancho))
        alto         = int(request.GET.get('alto',          evento.qr_alto))
        color_frente = request.GET.get('color_frente', evento.qr_color_frente)
        color_fondo  = request.GET.get('color_fondo',  evento.qr_color_fondo)
        fuente       = request.GET.get('fuente', evento.qr_fuente)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Parámetros inválidos'}, status=400)

    # Validar que la fuente sea una de las permitidas (seguridad)
    fuentes_permitidas = {f for f, _ in FUENTES_DISPONIBLES}
    if fuente not in fuentes_permitidas:
        fuente = 'Roboto-Bold.ttf'

    # Cargar fondo
    base_path = os.path.join(settings.BASE_DIR, 'cliente', 'static', 'img', 'asesor.jpeg')
    if evento.imagen_fondo:
        base_path = evento.imagen_fondo.path
    if not os.path.exists(base_path):
        return JsonResponse({'error': 'No hay imagen de fondo configurada'}, status=400)

    fondo = Image.open(base_path).convert("RGBA")
    entrada = fondo.copy()

    # Generar QR de muestra
    qr_obj = qrcode.QRCode(box_size=10, border=4)
    qr_obj.add_data("PREVIEW-QR")
    qr_obj.make(fit=True)

    transparente = (color_fondo == 'transparent')
    # back_color con alpha es ignorado por qrcode (genera en RGB); aplicar transparencia manualmente
    qr_img = qr_obj.make_image(fill_color=color_frente, back_color="white").convert("RGBA")
    pixels = list(qr_img.getdata())
    if transparente:
        pixels = [(r, g, b, 0) if r > 240 and g > 240 and b > 240 else (r, g, b, a)
                  for r, g, b, a in pixels]
    else:
        try:
            cb = color_fondo.lstrip('#')
            cb_rgb = (int(cb[0:2], 16), int(cb[2:4], 16), int(cb[4:6], 16))
            pixels = [(cb_rgb[0], cb_rgb[1], cb_rgb[2], 255) if r > 240 and g > 240 and b > 240
                      else (r, g, b, a) for r, g, b, a in pixels]
        except Exception:
            pass
    qr_img.putdata(pixels)
    qr_img = qr_img.resize((max(1, ancho), max(1, alto)), Image.Resampling.LANCZOS)
    capa = Image.new("RGBA", entrada.size, (0, 0, 0, 0))
    capa.paste(qr_img, (pos_x, pos_y), qr_img)
    entrada = Image.alpha_composite(entrada, capa)

    # Dibujar nombre de muestra
    draw = ImageDraw.Draw(entrada)
    nombre = "NOMBRE APELLIDO"
    font_path = os.path.join(settings.BASE_DIR, "cliente", "static", "fonts", fuente)
    font_size = 120
    while font_size > 40:
        try:
            font = ImageFont.truetype(font_path, font_size)
        except Exception:
            font = ImageFont.load_default()
            break
        bbox = draw.textbbox((0, 0), nombre, font=font)
        if (bbox[2] - bbox[0]) <= (ancho - 30):
            break
        font_size -= 8
    try:
        font = ImageFont.truetype(font_path, font_size)
    except Exception:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), nombre, font=font)
    texto_x = pos_x + (ancho // 2) - ((bbox[2] - bbox[0]) // 2)
    texto_y = pos_y + alto + 40
    draw.text((texto_x, texto_y), nombre, font=font, fill="white", stroke_width=6, stroke_fill="black")

    buffer = BytesIO()
    entrada.convert("RGB").save(buffer, format="JPEG", quality=85)
    img_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return JsonResponse({'imagen': img_b64, 'img_ancho': fondo.width, 'img_alto': fondo.height})


@login_required(login_url='/participantes/login/')
def preview_qr_ajax(request, evento_id):
    """Devuelve solo la imagen del QR (PNG base64) con los colores pedidos."""
    cf          = request.GET.get('color_frente', '#000000')
    cb          = request.GET.get('color_fondo',  '#ffffff')
    transparente = (cb == 'transparent')

    qr_obj = qrcode.QRCode(box_size=10, border=2)
    qr_obj.add_data("https://ede-evento.com/preview")
    qr_obj.make(fit=True)

    # qrcode siempre genera en modo RGB internamente; back_color con alpha es ignorado.
    # Generamos con fondo blanco y luego hacemos los píxeles blancos transparentes manualmente.
    qr_img = qr_obj.make_image(fill_color=cf, back_color="white").convert("RGBA")

    if transparente:
        pixels = list(qr_img.getdata())
        pixels = [(r, g, b, 0) if r > 240 and g > 240 and b > 240 else (r, g, b, a)
                  for r, g, b, a in pixels]
        qr_img.putdata(pixels)

    buffer = BytesIO()
    qr_img.save(buffer, format="PNG")
    img_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return JsonResponse({'qr': img_b64})


@login_required(login_url='/participantes/login/')
def guardar_config_qr(request, evento_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    evento = get_object_or_404(Evento, pk=evento_id)
    try:
        data = _json.loads(request.body)
        evento.qr_pos_x        = int(data['pos_x'])
        evento.qr_pos_y        = int(data['pos_y'])
        evento.qr_ancho        = int(data['ancho'])
        evento.qr_alto         = int(data['alto'])
        evento.qr_color_frente = data['color_frente']
        evento.qr_color_fondo  = data['color_fondo']
        fuentes_permitidas = {f for f, _ in FUENTES_DISPONIBLES}
        fuente = data.get('fuente', evento.qr_fuente)
        if fuente in fuentes_permitidas:
            evento.qr_fuente = fuente
        evento.save()
        return JsonResponse({'status': 'ok'})
    except (KeyError, ValueError) as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url='/participantes/login/')
def google_auth_inicio(request):
    """Redirige al usuario a la pantalla de autorización de Google."""
    if not getattr(settings, 'GOOGLE_CLIENT_ID', ''):
        messages.error(request, "❌ Google OAuth2 no está configurado aún. Pide al administrador que configure GOOGLE_CLIENT_ID.")
        return redirect('mi_cuenta')

    state = secrets.token_urlsafe(32)
    request.session['google_oauth_state'] = state

    params = {
        'client_id':     settings.GOOGLE_CLIENT_ID,
        'redirect_uri':  settings.GOOGLE_REDIRECT_URI,
        'response_type': 'code',
        'scope':         'https://www.googleapis.com/auth/gmail.send https://www.googleapis.com/auth/userinfo.email openid',
        'access_type':   'offline',
        'prompt':        'consent',   # Obligatorio para recibir refresh_token siempre
        'state':         state,
    }
    auth_url = 'https://accounts.google.com/o/oauth2/v2/auth?' + urllib.parse.urlencode(params)
    return redirect(auth_url)


@login_required(login_url='/participantes/login/')
def google_auth_callback(request):
    """Google llama aquí después de que el usuario autoriza. Guarda el refresh_token."""
    # Verificar CSRF state
    state_recibido = request.GET.get('state', '')
    state_guardado = request.session.pop('google_oauth_state', None)
    error          = request.GET.get('error')

    if error:
        messages.error(request, f"❌ Google rechazó la autorización: {error}")
        return redirect('mi_cuenta')

    if not state_guardado or state_recibido != state_guardado:
        messages.error(request, "❌ Error de seguridad en la autorización. Intenta conectar de nuevo.")
        return redirect('mi_cuenta')

    code = request.GET.get('code')
    if not code:
        messages.error(request, "❌ No se recibió el código de autorización de Google.")
        return redirect('mi_cuenta')

    # 1. Intercambiar código por tokens
    try:
        token_resp = requests.post(
            'https://oauth2.googleapis.com/token',
            data={
                'code':          code,
                'client_id':     settings.GOOGLE_CLIENT_ID,
                'client_secret': settings.GOOGLE_CLIENT_SECRET,
                'redirect_uri':  settings.GOOGLE_REDIRECT_URI,
                'grant_type':    'authorization_code',
            },
            timeout=15
        )
        token_data    = token_resp.json()
        refresh_token = token_data.get('refresh_token')
        access_token  = token_data.get('access_token')
    except Exception as e:
        logger.error(f"Error intercambiando código Google: {e}")
        messages.error(request, "❌ No se pudo comunicar con Google. Intenta de nuevo.")
        return redirect('mi_cuenta')

    if not refresh_token:
        messages.warning(
            request,
            "⚠️ Google no devolvió el refresh_token. Esto pasa cuando ya habías autorizado antes. "
            "Ve a myaccount.google.com → Seguridad → Aplicaciones con acceso → Elimina este app → Conecta de nuevo."
        )
        return redirect('mi_cuenta')

    # 2. Obtener email de la cuenta Google conectada
    try:
        userinfo_resp = requests.get(
            'https://www.googleapis.com/oauth2/v3/userinfo',
            headers={'Authorization': f'Bearer {access_token}'},
            timeout=10
        )
        google_email = userinfo_resp.json().get('email', '')
    except Exception:
        google_email = ''

    # 3. Guardar en el perfil del usuario
    perfil = get_object_or_404(PerfilUsuario, user=request.user)
    perfil.google_email         = google_email
    perfil.google_refresh_token = refresh_token
    perfil.save()

    messages.success(
        request,
        f"✅ Gmail '{google_email}' conectado correctamente. "
        f"Los correos de tus eventos se enviarán desde esa cuenta."
    )
    return redirect('mi_cuenta')


@login_required(login_url='/participantes/login/')
def google_desconectar(request):
    """Elimina el Gmail OAuth2 del perfil del usuario."""
    if request.method == 'POST':
        perfil = get_object_or_404(PerfilUsuario, user=request.user)
        email_anterior = perfil.google_email or ''
        perfil.google_email         = None
        perfil.google_refresh_token = None
        perfil.save()
        messages.success(request, f"Gmail '{email_anterior}' desconectado. Se usará el correo del sistema como respaldo.")
    return redirect('mi_cuenta')


@login_required(login_url='/participantes/login/')
@rol_requerido(['SUPERADMIN', 'ORGANIZADOR'])
def usuario_eliminar(request, pk):
    perfil_admin = get_object_or_404(PerfilUsuario, user=request.user)
    perfil = get_object_or_404(PerfilUsuario, pk=pk)
    username = perfil.user.username
    
    if perfil_admin.rol != 'SUPERADMIN':
        if perfil.rol != 'REGISTRADOR':
            messages.error(request, "No tienes permiso para eliminar este usuario.")
            return redirect('usuario_lista')
            
        mis_eventos = perfil_admin.eventos.all()
        shared_events = perfil.eventos.filter(id__in=mis_eventos.values_list('id', flat=True))
        if not shared_events.exists():
            messages.error(request, "No tienes permiso para eliminar este usuario.")
            return redirect('usuario_lista')
            
    if perfil.user == request.user:
        messages.error(request, "No puedes eliminar tu propio usuario actual.")
        return redirect('usuario_lista')
        
    perfil.user.delete()
    messages.success(request, f"Usuario '{username}' eliminado permanentemente.")
    return redirect('usuario_lista')
