from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.utils import timezone
from django.db.models import Count, Q
from .models import Plano, Lote, MovimientoLote, TipoCambio
from decimal import Decimal, InvalidOperation
from datetime import timedelta
import urllib.request
import json

# Nombre del grupo de Django para los asesores (solo pueden separar lotes).
GRUPO_ASESORES = "Asesores"

# Tipo de cambio
TC_FALLBACK     = Decimal("3.75")   # respaldo si nunca se pudo consultar
TC_MAX_EDAD_HRS = 3                 # refrescar si la cotización es más vieja que esto
TC_API_URL      = "https://open.er-api.com/v6/latest/USD"


def es_asesor(user):
    """True si el usuario pertenece al grupo Asesores (y no es staff)."""
    return user.groups.filter(name=GRUPO_ASESORES).exists()


def obtener_tipo_cambio():
    """
    Devuelve (usd_pen: Decimal, actualizado: datetime|None).
    Cachea en BD y solo consulta la API si la cotización está vieja.
    Si la API falla, usa el último valor conocido (o un respaldo).
    """
    tc = TipoCambio.objects.first()
    if tc and (timezone.now() - tc.actualizado) < timedelta(hours=TC_MAX_EDAD_HRS):
        return tc.usd_pen, tc.actualizado
    try:
        req = urllib.request.Request(TC_API_URL, headers={"User-Agent": "lotes/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        pen = Decimal(str(data["rates"]["PEN"])).quantize(Decimal("0.0001"))
        if not tc:
            tc = TipoCambio()
        tc.usd_pen = pen
        tc.save()
        return tc.usd_pen, tc.actualizado
    except Exception:
        if tc:
            return tc.usd_pen, tc.actualizado
        return TC_FALLBACK, None


# ──────────────────────────────────────────────────────────────
# API — tipo de cambio USD→PEN (público; lo usa el mapa para mostrar soles)
# URL: /lotes/tipo-cambio/
# ──────────────────────────────────────────────────────────────
def tipo_cambio(request):
    rate, actualizado = obtener_tipo_cambio()
    return JsonResponse({
        "usd_pen":     float(rate),
        "actualizado": timezone.localtime(actualizado).strftime("%d/%m/%Y %H:%M") if actualizado else None,
    })


def _get_lotes_data(plano):
    """Devuelve la lista de lotes del plano como JSON-serializable."""
    lotes = []
    for l in Lote.objects.filter(plano=plano).select_related("separado_por"):
        lotes.append({
            "id":     l.id,
            "numero": l.numero,
            "x":      l.x,
            "y":      l.y,
            "width":  l.width,
            "height": l.height,
            "estado": l.estado,
            "puntos": list(l.puntos) if l.puntos else None,
            "separado_por":    l.separado_por.username if l.separado_por_id else None,
            "separado_por_id": l.separado_por_id,
            "precio":          float(l.precio) if l.precio is not None else None,
            "separado_en":     timezone.localtime(l.separado_en).strftime("%d/%m/%Y %H:%M") if l.separado_en else None,
        })
    return lotes


# ──────────────────────────────────────────────────────────────
# VISTA ADMIN — solo staff, puede modificar estados
# URL: /plano/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def ver_plano(request):
    asesor = es_asesor(request.user)
    # Staff = control total. Asesor = solo separar. Cualquier otro = mapa público.
    if not request.user.is_staff and not asesor:
        return redirect('mapa_publico')

    plano = Plano.objects.first()
    lotes = _get_lotes_data(plano) if plano else []

    # solo_separar = True únicamente para asesores que NO son staff
    solo_separar = asesor and not request.user.is_staff

    return render(request, "lotes/plano_admin.html", {
        "plano": plano,
        "lotes": json.dumps(lotes),
        "solo_separar": solo_separar,
        "usuario_actual": request.user.username,
        "es_admin": request.user.is_staff,
    })


# ──────────────────────────────────────────────────────────────
# VISTA PÚBLICA — sin login, solo lectura
# URL: /mapa/
# ──────────────────────────────────────────────────────────────
def ver_mapa_publico(request):
    plano = Plano.objects.first()
    lotes = _get_lotes_data(plano) if plano else []

    # Estadísticas para el panel de resumen
    # Los lotes "bloqueados" no son lotes reales → se excluyen del conteo
    vendidos    = sum(1 for l in lotes if l['estado'] == 'vendido')
    reservados  = sum(1 for l in lotes if l['estado'] == 'reservado')
    bloqueados  = sum(1 for l in lotes if l['estado'] == 'bloqueado')
    total       = len(lotes) - bloqueados
    disponibles = total - vendidos - reservados

    return render(request, "lotes/plano_publico.html", {
        "plano":       plano,
        "lotes":       json.dumps(lotes),
        "total":       total,
        "vendidos":    vendidos,
        "reservados":  reservados,
        "disponibles": disponibles,
    })


# ──────────────────────────────────────────────────────────────
# API — cambiar estado (solo staff)
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def cambiar_estado_lote(request):
    if request.method == "POST":
        asesor = es_asesor(request.user)
        if not request.user.is_staff and not asesor:
            return JsonResponse({"error": "No autorizado"}, status=403)

        data         = json.loads(request.body)
        lote         = get_object_or_404(Lote, id=data["id"])
        nuevo_estado = data.get("estado")

        if nuevo_estado not in ["disponible", "vendido", "reservado", "bloqueado"]:
            return JsonResponse({"error": "Estado inválido"}, status=400)

        es_admin = request.user.is_staff

        # ── Reglas para ASESORES (el admin no tiene restricciones) ──
        if not es_admin:
            if nuevo_estado == "reservado":
                # No se separan zonas bloqueadas (no son lotes) ni lotes vendidos
                if lote.estado == "bloqueado":
                    return JsonResponse({"error": "Ese espacio no es un lote."}, status=403)
                if lote.estado == "vendido":
                    return JsonResponse({"error": "Ese lote ya está vendido."}, status=403)
                # No puede robar una separación activa de otro asesor
                if (lote.estado == "reservado" and lote.separado_por_id
                        and lote.separado_por_id != request.user.id):
                    return JsonResponse(
                        {"error": "Ese lote ya fue separado por otro asesor."}, status=403)
            elif nuevo_estado == "disponible":
                # Solo puede desechar lo que él mismo separó
                if lote.separado_por_id != request.user.id:
                    return JsonResponse(
                        {"error": "Solo puedes quitar las separaciones que tú hiciste."}, status=403)
            else:
                return JsonResponse(
                    {"error": "Solo puedes separar o quitar tus separaciones."}, status=403)

        # ── Aplicar el cambio ──
        ahora = timezone.now()
        lote.estado = nuevo_estado
        if nuevo_estado == "reservado":
            lote.separado_por = request.user
            lote.separado_en  = ahora
        elif nuevo_estado == "disponible":
            lote.separado_por = None
            lote.separado_en  = None
            lote.vendido_en   = None
        elif nuevo_estado == "vendido":
            lote.vendido_en   = ahora
        elif nuevo_estado == "bloqueado":
            lote.precio       = None   # las zonas bloqueadas no llevan precio
        lote.save()

        # Historial simple: quién hizo qué y cuándo
        MovimientoLote.objects.create(lote=lote, usuario=request.user, estado=nuevo_estado)

        return JsonResponse({
            "status":       "ok",
            "estado":       lote.estado,
            "separado_por": lote.separado_por.username if lote.separado_por_id else None,
            "separado_en":  timezone.localtime(lote.separado_en).strftime("%d/%m/%Y %H:%M") if lote.separado_en else None,
            "precio":       float(lote.precio) if lote.precio is not None else None,
        })
    return JsonResponse({"error": "Método no permitido"}, status=405)


# ──────────────────────────────────────────────────────────────
# API — estados de lotes (para actualización en vivo / polling)
# Devuelve solo lo necesario, ligero. Cualquiera con acceso al mapa.
# URL: /lotes/estados/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def estados_lotes(request):
    if not request.user.is_staff and not es_asesor(request.user):
        return JsonResponse({"error": "No autorizado"}, status=403)
    plano = Plano.objects.first()
    datos = []
    if plano:
        for l in Lote.objects.filter(plano=plano).select_related("separado_por"):
            datos.append({
                "id":           l.id,
                "numero":       l.numero,
                "estado":       l.estado,
                "separado_por": l.separado_por.username if l.separado_por_id else None,
                "precio":       float(l.precio) if l.precio is not None else None,
            })
    return JsonResponse({"lotes": datos})


# ──────────────────────────────────────────────────────────────
# API — fijar precio de un lote (solo staff)
# URL: /lotes/set-precio/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def set_precio_lote(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    if not request.user.is_staff:
        return JsonResponse({"error": "No autorizado"}, status=403)

    data    = json.loads(request.body)
    lote    = get_object_or_404(Lote, id=data["id"])
    if lote.estado == "bloqueado":
        return JsonResponse({"error": "Las zonas bloqueadas no llevan precio."}, status=400)
    monto   = data.get("monto", data.get("precio"))    # compat
    moneda  = (data.get("moneda") or "USD").upper()
    rate, _ = obtener_tipo_cambio()

    if monto in (None, "", "null"):
        lote.precio = None
    else:
        try:
            monto = Decimal(str(monto))
        except (InvalidOperation, TypeError, ValueError):
            return JsonResponse({"error": "Precio inválido"}, status=400)
        if monto < 0:
            return JsonResponse({"error": "Precio inválido"}, status=400)
        # Se guarda SIEMPRE en dólares (base estable). Si lo ponen en soles, se convierte.
        usd = (monto / rate) if moneda == "PEN" else monto
        lote.precio = usd.quantize(Decimal("0.01"))

    lote.save(update_fields=["precio"])
    return JsonResponse({
        "status":  "ok",
        "precio":  float(lote.precio) if lote.precio is not None else None,  # en USD
        "usd_pen": float(rate),
    })


# ──────────────────────────────────────────────────────────────
# API — fijar/corregir el NÚMERO de un lote (solo staff)
# URL: /lotes/set-numero/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def set_numero_lote(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    if not request.user.is_staff:
        return JsonResponse({"error": "No autorizado"}, status=403)
    data   = json.loads(request.body)
    lote   = get_object_or_404(Lote, id=data["id"])
    numero = (data.get("numero") or "").strip()
    lote.numero = numero or None
    lote.save(update_fields=["numero"])
    return JsonResponse({"status": "ok", "numero": lote.numero})


# ──────────────────────────────────────────────────────────────
# API — fijar precio MASIVO a todos los lotes (solo staff)
# URL: /lotes/set-precio-masivo/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def set_precio_masivo(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    if not request.user.is_staff:
        return JsonResponse({"error": "No autorizado"}, status=403)

    data           = json.loads(request.body)
    monto          = data.get("monto")
    moneda         = (data.get("moneda") or "USD").upper()
    solo_sin_precio = bool(data.get("solo_sin_precio"))
    rate, _        = obtener_tipo_cambio()

    # Las zonas bloqueadas no son lotes → nunca reciben precio
    qs = Lote.objects.exclude(estado="bloqueado")
    if solo_sin_precio:
        qs = qs.filter(precio__isnull=True)

    if monto in (None, "", "null"):
        precio_usd = None   # vaciar precio
    else:
        try:
            monto = Decimal(str(monto))
        except (InvalidOperation, TypeError, ValueError):
            return JsonResponse({"error": "Precio inválido"}, status=400)
        if monto < 0:
            return JsonResponse({"error": "Precio inválido"}, status=400)
        precio_usd = ((monto / rate) if moneda == "PEN" else monto).quantize(Decimal("0.01"))

    actualizados = qs.update(precio=precio_usd)
    return JsonResponse({
        "status":       "ok",
        "actualizados": actualizados,
        "precio":       float(precio_usd) if precio_usd is not None else None,  # USD
        "usd_pen":      float(rate),
    })


# ──────────────────────────────────────────────────────────────
# REPORTE por asesor (solo staff)
# URL: /lotes/reporte/
# ──────────────────────────────────────────────────────────────
def _resumen_asesores():
    """Agrega lotes por asesor: separados (reservados) y vendidos."""
    filas = (
        User.objects
        .filter(lotes_separados__isnull=False)
        .annotate(
            separados=Count("lotes_separados", filter=Q(lotes_separados__estado="reservado"), distinct=True),
            vendidos =Count("lotes_separados", filter=Q(lotes_separados__estado="vendido"),   distinct=True),
        )
        .order_by("-vendidos", "-separados", "username")
        .distinct()
    )
    return filas


@login_required(login_url='/participantes/login/')
def reporte_asesores(request):
    if not request.user.is_staff:
        return redirect('mapa_publico')

    filas = _resumen_asesores()
    total_separados = sum(f.separados for f in filas)
    total_vendidos  = sum(f.vendidos  for f in filas)

    return render(request, "lotes/reporte_asesores.html", {
        "filas":           filas,
        "total_separados": total_separados,
        "total_vendidos":  total_vendidos,
    })


# ──────────────────────────────────────────────────────────────
# EXPORTAR — reporte por asesor a Excel (solo staff)
# URL: /lotes/reporte/excel/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def reporte_asesores_excel(request):
    if not request.user.is_staff:
        return redirect('mapa_publico')

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte por asesor"

    encabezados = ["Asesor", "Separados (activos)", "Vendidos", "Total gestionados"]
    ws.append(encabezados)
    azul = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul

    for f in _resumen_asesores():
        ws.append([f.username, f.separados, f.vendidos, f.separados + f.vendidos])

    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = ancho + 4

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="reporte_asesores.xlsx"'
    wb.save(resp)
    return resp


# ──────────────────────────────────────────────────────────────
# EXPORTAR — respaldo de TODOS los estados de lotes a Excel (solo staff)
# URL: /lotes/exportar-estados/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def exportar_estados_excel(request):
    if not request.user.is_staff:
        return redirect('mapa_publico')

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estados de lotes"

    rate, _ = obtener_tipo_cambio()
    ws.append(["Lote N°", "ID", "Estado", "Precio (US$)", f"Precio (S/ · TC {rate})",
               "Separado por", "Separado el", "Vendido el"])
    azul = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul

    fmt = lambda dt: timezone.localtime(dt).strftime("%d/%m/%Y %H:%M") if dt else ""

    def _num_key(l):
        try:
            return (0, int(l.numero))
        except (TypeError, ValueError):
            return (1, l.id)   # los sin número, al final

    # Las zonas bloqueadas no son lotes → se excluyen del respaldo. Orden por número de lote.
    qs = Lote.objects.select_related("separado_por").exclude(estado="bloqueado")
    for l in sorted(qs, key=_num_key):
        usd = float(l.precio) if l.precio is not None else ""
        pen = round(float(l.precio) * float(rate), 2) if l.precio is not None else ""
        ws.append([
            l.numero or "", l.id, l.estado, usd, pen,
            l.separado_por.username if l.separado_por_id else "",
            fmt(l.separado_en), fmt(l.vendido_en),
        ])

    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = ancho + 3

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fecha = timezone.localtime().strftime("%Y%m%d_%H%M")
    resp["Content-Disposition"] = f'attachment; filename="respaldo_lotes_{fecha}.xlsx"'
    wb.save(resp)
    return resp


# ──────────────────────────────────────────────────────────────
# CREAR ASESOR — solo staff. Crea un usuario que solo puede separar.
# URL: /lotes/crear-asesor/
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def crear_asesor(request):
    # Solo el administrador (staff) puede crear asesores
    if not request.user.is_staff:
        return redirect('mapa_publico')

    creado = None
    error  = None

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()

        if not username or not password:
            error = "El usuario y la contraseña son obligatorios."
        elif User.objects.filter(username=username).exists():
            error = "Ese nombre de usuario ya existe. Elige otro."
        else:
            # Usuario normal (NO staff, NO superuser) → no toca el sistema actual
            user = User.objects.create_user(username=username, password=password)
            user.is_staff = False
            user.save()
            # Lo metemos al grupo Asesores (se crea solo si no existe)
            grupo, _ = Group.objects.get_or_create(name=GRUPO_ASESORES)
            user.groups.add(grupo)
            creado = username

    # Lista de asesores existentes para mostrarlos
    asesores = User.objects.filter(groups__name=GRUPO_ASESORES).order_by("username")

    return render(request, "lotes/crear_asesor.html", {
        "creado":   creado,
        "error":    error,
        "asesores": asesores,
    })


# ──────────────────────────────────────────────────────────────
# API — guardar lote (solo staff)
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def guardar_lote(request):
    if request.method == "POST":
        if not request.user.is_staff:
            return JsonResponse({"error": "No autorizado"}, status=403)
        data  = json.loads(request.body)
        plano = get_object_or_404(Plano, id=data["plano_id"])
        lote  = Lote.objects.create(
            plano=plano,
            puntos=data.get("puntos"),
            x=data.get("x"), y=data.get("y"),
            width=data.get("width"), height=data.get("height"),
            estado=data["estado"]
        )
        return JsonResponse({"status": "ok", "id": lote.id})


# ──────────────────────────────────────────────────────────────
# API — eliminar lote (solo staff)
# ──────────────────────────────────────────────────────────────
@login_required(login_url='/participantes/login/')
def eliminar_lote(request):
    if request.method == "POST":
        if not request.user.is_staff:
            return JsonResponse({"error": "No autorizado"}, status=403)
        data = json.loads(request.body)
        lote = get_object_or_404(Lote, id=data["id"])
        lote.delete()
        return JsonResponse({"status": "ok"})
    return JsonResponse({"error": "Método no permitido"}, status=405)
